"""事业编导入链路测试."""
import os
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

os.environ.setdefault("DATABASE_URL", "postgresql+asyncpg://test:test@localhost/test")
os.environ.setdefault("SECRET_KEY", "test-secret")

from app.services.position_smart_import_service import PositionSmartImportService
from app.services.shiye_import_service import ShiyeImportService, TOTAL_TABLE_LAYOUT


class FakeScalarResult:
    def __init__(self, value):
        self._value = value

    def scalar_one_or_none(self):
        return self._value


class FakeImportDB:
    def __init__(self):
        self.added = []
        self.committed = False

    async def execute(self, _statement):
        return FakeScalarResult(None)

    def add(self, item):
        self.added.append(item)

    async def flush(self):
        return None

    async def commit(self):
        self.committed = True


def build_total_table_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "2025年江苏事业单位"
    ws.append([
        "地市", "区县", "主管部门", "招录单位", "岗位代码", "岗位名称", "岗位说明", "笔试类别",
        "岗位等级", "经费来源", "招录人数", "学历", "学位", "专业要求", "其他条件", "招聘对象",
        "开考比例", "笔面试占比", "面试比例", "备注", "报名人数", "竞争比", "进面最低分", "进面最高分",
    ])
    ws.append([
        "南京市", "南京市属", "南京市教育局", "南京市第一中学", "001", "教师岗", "承担语文教学", "专技类",
        "专技十二级", "全额拨款", 2, "本科及以上", "学士", "汉语言文学", "取得相应学位", "2025年毕业生",
        "1:3", "笔试50%，面试50%", "3:1", "编内", 80, 40, 72.5, 78.0,
    ])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_multi_sheet_shiye_workbook() -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "省属"
    ws2 = wb.create_sheet("南京")

    for ws, unit_name, title, code in (
        (ws1, "机关后勤服务中心", "综合文秘", "001"),
        (ws2, "南京市第一中学", "教师岗", "002"),
    ):
        ws.append(["2025年江苏事业单位统一公开招聘岗位表"])
        ws.append([
            "部门名称", "主管部门代码", "招聘单位", None, None, "招聘岗位", "岗位代码", "岗位类别",
            "招聘人数", "开考比例", "招聘条件", None, None, None, "招聘部门（单位）考试形式和所占比例",
            "其他说明", "政策咨询电话及联系人", "最高分", "最低分", "地区",
        ])
        ws.append([
            None, None, "单位名称", "单位代码", "经费来源", "岗位名称", None, None,
            None, None, "学历", "专业", "其他条件", "招聘对象", None, None, None, "最高分", "最低分", "地区",
        ])
        ws.append([
            "教育局", None, unit_name, "1001", "全额拨款", title, code, "管理类",
            1, "1:3", "本科及以上", "汉语言文学", "取得相应学位", "社会人员", "笔试50%，面试50%",
            "编内", "025-12345678", 78.0, 72.0, "市直属",
        ])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@pytest.mark.unit
class TestShiyeImportServices:
    def test_smart_import_parses_total_table_fields(self):
        content = build_total_table_workbook()

        result = PositionSmartImportService.parse_file(content)

        assert result["type"] == "complete"
        assert result["row_count"] == 1
        row = result["rows"][0]
        assert row["department"] == "南京市第一中学"
        assert row["title"] == "教师岗"
        assert row["position_code"] == "001"
        assert row["funding_source"] == "全额拨款"
        assert row["recruitment_target"] == "2025年毕业生"
        assert row["apply_count"] == 80
        assert row["competition_ratio"] == 40.0

    def test_smart_import_parses_multi_sheet_shiye_headers(self):
        content = build_multi_sheet_shiye_workbook()

        result = PositionSmartImportService.parse_file(content)

        assert result["type"] == "position"
        assert result["row_count"] == 2
        titles = {row["title"] for row in result["rows"]}
        assert titles == {"综合文秘", "教师岗"}
        assert all(row["funding_source"] == "全额拨款" for row in result["rows"])
        assert all(row["recruitment_target"] == "社会人员" for row in result["rows"])

    @pytest.mark.asyncio
    async def test_shiye_import_service_supports_total_table_layout(self, tmp_path):
        content = build_total_table_workbook()
        file_path = tmp_path / "2025年江苏事业单位总表.xlsx"
        file_path.write_bytes(content)
        db = FakeImportDB()
        ws = load_workbook(file_path, data_only=True, read_only=True).active

        result = await ShiyeImportService.import_file(db, str(file_path), 2025)

        assert ShiyeImportService.detect_layout(ws) == TOTAL_TABLE_LAYOUT
        assert result == {"year": 2025, "inserted": 1, "updated": 0, "skipped": 0}
        assert db.committed is True
        position = db.added[0]
        assert position.funding_source == "全额拨款"
        assert position.recruitment_target == "2025年毕业生"
        assert position.degree == "学士"
        assert position.apply_count == 80
        assert position.exam_weight_ratio == "笔试50%，面试50%"
