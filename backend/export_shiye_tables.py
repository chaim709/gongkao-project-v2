"""导出2023/2024/2025年江苏事业单位总表为Excel"""
import asyncio
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from app.database import AsyncSessionLocal
from app.models.position import Position

# 统一列定义：(DB字段, 表头名, 列宽)
COLUMNS = [
    ('city', '地市', 10),
    ('location', '区县', 12),
    ('supervising_dept', '主管部门', 22),
    ('department', '招录单位', 25),
    ('position_code', '岗位代码', 10),
    ('title', '岗位名称', 18),
    ('description', '岗位说明', 25),
    ('exam_category', '笔试类别', 14),
    ('position_level', '岗位等级', 12),
    ('funding_source', '经费来源', 12),
    ('recruitment_count', '招录人数', 8),
    ('education', '学历', 14),
    ('degree', '学位', 12),
    ('major', '专业要求', 30),
    ('other_requirements', '其他条件', 25),
    ('recruitment_target', '招聘对象', 12),
    ('exam_ratio', '开考比例', 10),
    ('exam_weight_ratio', '笔面试占比', 20),
    ('interview_ratio', '面试比例', 12),
    ('remark', '备注', 20),
    ('apply_count', '报名人数', 10),
    ('competition_ratio', '竞争比', 8),
    ('min_interview_score', '进面最低分', 10),
    ('max_interview_score', '进面最高分', 10),
]

HEADER_FONT = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGN = Alignment(vertical='center', wrap_text=False)
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
EVEN_FILL = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')


async def export_year(year: int):
    async with AsyncSessionLocal() as db:
        positions = (await db.execute(
            select(Position).where(
                Position.year == year, Position.exam_type == '事业单位'
            ).order_by(Position.city, Position.department, Position.position_code)
        )).scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{year}年江苏事业单位'

    # 写表头
    for ci, (field, header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'
    # 自动筛选
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNS))}1'

    # 写数据
    for ri, p in enumerate(positions, 2):
        for ci, (field, _, _) in enumerate(COLUMNS, 1):
            val = getattr(p, field, None)
            if val is not None:
                # 清理换行
                if isinstance(val, str):
                    val = val.replace('\n', ' ').replace('\r', '')
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER
            if ri % 2 == 0:
                cell.fill = EVEN_FILL

    out = f'/Users/chaim/Desktop/{year}年江苏事业单位总表.xlsx'
    wb.save(out)
    wb.close()
    print(f'{year}年: {len(positions)} 条 → {out}')


async def main():
    for year in (2023, 2024, 2025):
        await export_year(year)


if __name__ == '__main__':
    asyncio.run(main())
