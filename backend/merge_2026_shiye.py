"""整合2026年江苏省各地市事业单位岗位表为统一总表

处理72个xlsx/xls文件，自动识别表头，统一字段格式
"""
import os
import re
import openpyxl
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 标准字段映射（目标字段 → 可能的源字段名）
FIELD_MAP = {
    'city': ['地市', '市', '城市'],
    'location': ['区县', '县区', '区、县级市', '县（市、区）'],
    'supervising_dept': ['主管部门', '主管单位'],
    'department': ['招聘单位', '招录单位', '单位名称', '用人单位'],
    'position_code': ['岗位代码', '职位代码', '岗位编号'],
    'title': ['岗位名称', '职位名称', '岗位'],
    'description': ['岗位说明', '岗位简介', '岗位描述', '职位简介'],
    'exam_category': ['笔试类别', '笔试科目', '考试类别'],
    'position_level': ['岗位等级', '职级'],
    'funding_source': ['经费来源', '单位经费来源'],
    'recruitment_count': ['招聘人数', '招录人数', '招聘\n人数', '招录\n人数', '人数'],
    'education': ['学历', '学历要求'],
    'degree': ['学位', '学位要求'],
    'major': ['专业', '专业要求'],
    'other_requirements': ['其他条件', '其他要求', '备注'],
    'recruitment_target': ['招聘对象', '招录对象'],
    'exam_ratio': ['开考比例'],
}

# 反向映射：源字段名 → 目标字段
REVERSE_MAP = {}
for target, sources in FIELD_MAP.items():
    for src in sources:
        REVERSE_MAP[src.replace('\n', '').strip()] = target


def normalize_header(h):
    """标准化表头名称"""
    if not h:
        return ''
    return str(h).replace('\n', '').replace(' ', '').strip()


def find_header_row(rows, max_check=10):
    """查找表头行（包含"岗位"或"招聘"等关键词）"""
    for i, row in enumerate(rows[:max_check]):
        row_text = ''.join([str(c or '') for c in row])
        if any(kw in row_text for kw in ['岗位', '招聘', '单位', '专业', '学历']):
            # 检查是否有足够多的非空列
            non_empty = sum(1 for c in row if c and str(c).strip())
            if non_empty >= 5:
                return i
    return 0


def parse_xlsx(fp):
    """解析xlsx文件"""
    wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return all_rows


def parse_xls(fp):
    """解析xls文件"""
    wb = xlrd.open_workbook(fp)
    ws = wb.sheet_by_index(0)
    all_rows = [ws.row_values(i) for i in range(ws.nrows)]
    return all_rows


def extract_city_from_filename(fname):
    """从文件名提取地市"""
    # 匹配常见地市名
    cities = ['南京', '苏州', '无锡', '常州', '南通', '徐州', '扬州', '镇江', '盐城', '泰州', '淮安', '连云港', '宿迁']
    for city in cities:
        if city in fname:
            return city + '市'
    # 匹配县级市
    if '沭阳' in fname:
        return '宿迁市'
    if '泗阳' in fname:
        return '宿迁市'
    if '仪征' in fname:
        return '扬州市'
    if '高邮' in fname:
        return '扬州市'
    if '宝应' in fname:
        return '扬州市'
    if '句容' in fname or '丹阳' in fname:
        return '镇江市'
    if '宿豫' in fname:
        return '宿迁市'
    return '未知'


def parse_file(fp, fname):
    """解析单个文件，返回标准化数据"""
    try:
        if fp.endswith('.xlsx'):
            all_rows = parse_xlsx(fp)
        else:
            all_rows = parse_xls(fp)
    except Exception as e:
        print(f'  ✗ {fname}: {e}')
        return []

    # 查找表头
    header_idx = find_header_row(all_rows)
    if header_idx >= len(all_rows):
        print(f'  ✗ {fname}: 未找到表头')
        return []

    headers = [normalize_header(h) for h in all_rows[header_idx]]

    # 映射列索引
    col_map = {}
    for ci, h in enumerate(headers):
        if h in REVERSE_MAP:
            col_map[REVERSE_MAP[h]] = ci

    if not col_map:
        print(f'  ✗ {fname}: 无法识别字段')
        return []

    # 提取数据
    city_from_file = extract_city_from_filename(fname)
    results = []

    for row in all_rows[header_idx + 1:]:
        # 跳过空行
        if not any(row):
            continue

        data = {}
        for field, ci in col_map.items():
            val = row[ci] if ci < len(row) else None
            if val is not None and str(val).strip():
                data[field] = str(val).strip()

        # 必须有单位或岗位
        if not data.get('department') and not data.get('title'):
            continue

        # 补充地市
        if 'city' not in data:
            data['city'] = city_from_file

        results.append(data)

    print(f'  ✓ {fname}: {len(results)} 条')
    return results


def main():
    dir_path = '/Users/chaim/attachments/江苏省岗位表集合/'
    files = sorted([f for f in os.listdir(dir_path) if f.endswith(('.xlsx', '.xls'))])

    print(f'开始处理 {len(files)} 个文件...\n')

    all_data = []
    for fname in files:
        fp = os.path.join(dir_path, fname)
        data = parse_file(fp, fname)
        all_data.extend(data)

    print(f'\n总计: {len(all_data)} 条岗位')

    # 导出Excel
    wb = Workbook()
    ws = wb.active
    ws.title = '2026年江苏事业单位'

    # 定义列
    columns = [
        ('city', '地市', 10),
        ('location', '区县', 12),
        ('supervising_dept', '主管部门', 22),
        ('department', '招录单位', 25),
        ('position_code', '岗位代码', 10),
        ('title', '岗位名称', 18),
        ('description', '岗位说明', 25),
        ('exam_category', '笔试类别', 14),
        ('position_level', '岗位等级', 12),
        ('funding_source', '经费来源', 12),
        ('recruitment_count', '招聘人数', 8),
        ('education', '学历', 14),
        ('degree', '学位', 12),
        ('major', '专业要求', 30),
        ('other_requirements', '其他条件', 25),
        ('recruitment_target', '招聘对象', 12),
        ('exam_ratio', '开考比例', 10),
    ]

    # 样式
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    cell_align = Alignment(vertical='center')
    border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    # 写表头
    for ci, (field, header, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(columns))}1'

    # 写数据
    for ri, item in enumerate(all_data, 2):
        for ci, (field, _, _) in enumerate(columns, 1):
            val = item.get(field, '')
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = cell_align
            cell.border = border

    out = '/Users/chaim/Desktop/2026年江苏事业单位总表.xlsx'
    wb.save(out)
    wb.close()
    print(f'\n已导出: {out}')


if __name__ == '__main__':
    main()
