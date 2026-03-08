"""合并报名人数数据到 positions 表

报名数据格式分3类：
A. 标准垂直(多数城市): 城市,部门名称[code],职位名称[code],...,报名成功人数
B. 苏州特殊列: 招聘单位[code],招录人数,区县,岗位[code],...,报名成功数
C. 水平排列(泰州/淮安/徐州): 多城市并排，每块5-6列

匹配策略：
1. 优先: (city, department_code, position_code)
2. 备选: (city, department子串, title)
"""
import asyncio
import re
import openpyxl
from sqlalchemy import select
from app.database import AsyncSessionLocal
from app.models.position import Position


CODE_PATTERN = re.compile(r'\[([^\]]+)\]\s*$')


def extract_code(text: str) -> tuple[str, str]:
    """从 '单位名称 [code]' 中提取名称和代码"""
    if not text:
        return '', ''
    text = str(text).strip()
    m = CODE_PATTERN.search(text)
    if m:
        code = m.group(1).strip()
        name = text[:m.start()].strip()
        # 去掉 "主管部门-" 前缀
        if '-' in name:
            name = name.split('-', 1)[1].strip()
        return name, code
    return text, ''


def parse_standard_sheet(ws, sheet_name: str, data_start: int, dept_col: int,
                         pos_col: int, apply_col: int) -> list[dict]:
    """解析标准垂直格式的 Sheet"""
    rows = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row or not row[dept_col]:
            continue
        dept_name, dept_code = extract_code(row[dept_col])
        pos_name, pos_code = extract_code(row[pos_col])
        try:
            apply_count = int(row[apply_col]) if row[apply_col] else None
        except (ValueError, TypeError):
            apply_count = None
        if dept_name and pos_name and apply_count is not None:
            rows.append({
                'dept_name': dept_name, 'dept_code': dept_code,
                'pos_name': pos_name, 'pos_code': pos_code,
                'apply_count': apply_count,
            })
    return rows


def find_block_boundaries(ws, header_row: int) -> list[tuple[int, int]]:
    """动态检测水平排列的 block 边界，通过找空列分隔"""
    row = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    blocks = []
    block_start = None
    for ci in range(len(row)):
        is_empty = row[ci] is None or str(row[ci]).strip() == ''
        if not is_empty and block_start is None:
            block_start = ci
        elif is_empty and block_start is not None:
            blocks.append((block_start, ci))
            block_start = None
    if block_start is not None:
        blocks.append((block_start, len(row)))
    return blocks


def parse_horizontal_sheet(ws, data_start: int, header_row: int,
                           has_city_col: bool = False) -> list[dict]:
    """解析水平排列格式 (泰州/淮安/徐州)
    动态检测 block 边界，支持不等宽 block
    """
    blocks = find_block_boundaries(ws, header_row)
    rows = []

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row:
            continue
        for (bs, be) in blocks:
            # 跳过城市列（泰州有，淮安/徐州没有）
            offset = 1 if has_city_col else 0
            dept_idx = bs + offset
            pos_idx = bs + offset + 1
            # 报名成功人数 = block 最后一列（空列前一列）
            apply_idx = be - 1

            if dept_idx >= len(row) or pos_idx >= len(row):
                continue

            dept_val = row[dept_idx]
            pos_val = row[pos_idx]
            if not dept_val or not pos_val:
                continue

            dept_name, dept_code = extract_code(dept_val)
            pos_name, pos_code = extract_code(pos_val)

            apply_val = row[apply_idx] if apply_idx < len(row) else None
            try:
                apply_count = int(apply_val) if apply_val else None
            except (ValueError, TypeError):
                apply_count = None

            if dept_name and pos_name and apply_count is not None:
                rows.append({
                    'dept_name': dept_name, 'dept_code': dept_code,
                    'pos_name': pos_name, 'pos_code': pos_code,
                    'apply_count': apply_count,
                })
    return rows


# Sheet名 → city标准名
SHEET_CITY_MAP = {
    '省属': '省属', '南京': '南京市', '苏州': '苏州市', '无锡': '无锡市',
    '常州': '常州市', '南通': '南通市', '扬州': '扬州市', '镇江': '镇江市',
    '泰州': '泰州市', '盐城': '盐城市', '连云港': '连云港市',
    '淮安': '淮安市', '徐州': '徐州市', '宿迁': '宿迁市',
}


def parse_all_sheets(wb) -> dict[str, list[dict]]:
    """解析所有Sheet，返回 {city: [rows]}"""
    result = {}

    for raw_name in wb.sheetnames:
        short_name = raw_name.split('（')[0].strip()
        city = SHEET_CITY_MAP.get(short_name)
        if not city:
            print(f'  跳过未知Sheet: {raw_name}')
            continue
        ws = wb[raw_name]

        if short_name == '苏州':
            # Format B: 招聘单位[code], 招录人数, 区县, 岗位[code], 报名人数, 审核通过数, 报名成功数
            rows = parse_standard_sheet(ws, raw_name, 2, dept_col=0, pos_col=3, apply_col=6)
        elif short_name == '泰州':
            # 水平排列，有城市列，header在行1
            rows = parse_horizontal_sheet(ws, data_start=3, header_row=1, has_city_col=True)
        elif short_name in ('淮安', '徐州'):
            # 水平排列，无城市列，header在行2
            rows = parse_horizontal_sheet(ws, data_start=4, header_row=2, has_city_col=False)
        elif short_name == '南通':
            # 南通 data_start=3 (行1标题, 行2表头, 行3开始)
            rows = parse_standard_sheet(ws, raw_name, 3, dept_col=1, pos_col=2, apply_col=5)
        elif short_name == '宿迁':
            # 宿迁无空行2, data_start=2
            rows = parse_standard_sheet(ws, raw_name, 2, dept_col=1, pos_col=2, apply_col=5)
        elif short_name in ('南京', '镇江'):
            # 8列格式, 报名成功人数在 col 7
            rows = parse_standard_sheet(ws, raw_name, 3, dept_col=1, pos_col=2, apply_col=7)
        elif short_name == '盐城':
            # 6列但有开考比例和招考人数, 报名成功人数在 col 5
            rows = parse_standard_sheet(ws, raw_name, 3, dept_col=1, pos_col=2, apply_col=5)
        else:
            # Format A 标准: 城市, 部门名称, 职位名称, 审核通过, 未审核, 报名成功人数
            rows = parse_standard_sheet(ws, raw_name, 3, dept_col=1, pos_col=2, apply_col=5)

        result[city] = rows
        print(f'  {short_name:6s}: 解析 {len(rows)} 条报名数据')

    return result


async def merge_apply_count(apply_data: dict[str, list[dict]]):
    """将报名人数合并到 positions 表"""
    async with AsyncSessionLocal() as db:
        total_matched = 0
        total_unmatched = 0

        for city, rows in apply_data.items():
            # 预加载该城市所有岗位
            positions = (await db.execute(
                select(Position).where(
                    Position.year == 2025,
                    Position.exam_type == '事业单位',
                    Position.city == city,
                )
            )).scalars().all()

            # 建索引: (dept_code, pos_code) → [positions]
            by_code = {}
            for p in positions:
                key = (p.department_code or '', p.position_code or '')
                by_code.setdefault(key, []).append(p)

            # 建索引: (dept_name_substr, title) → [positions]
            by_name = {}
            for p in positions:
                key = (p.department or '', p.title or '')
                by_name.setdefault(key, []).append(p)

            matched = 0
            unmatched = 0
            matched_ids = set()

            for row in rows:
                position = None

                # 策略1: 按 (dept_code, pos_code) 匹配
                if row['dept_code'] and row['pos_code']:
                    candidates = by_code.get((row['dept_code'], row['pos_code']), [])
                    for c in candidates:
                        if c.id not in matched_ids:
                            position = c
                            break

                # 策略2: 按名称匹配
                if not position:
                    # 精确匹配
                    candidates = by_name.get((row['dept_name'], row['pos_name']), [])
                    for c in candidates:
                        if c.id not in matched_ids:
                            position = c
                            break

                # 策���3: 模糊匹配 - dept_name 包含
                if not position:
                    for p in positions:
                        if p.id in matched_ids:
                            continue
                        dept_match = (row['dept_name'] in (p.department or '')
                                     or (p.department or '') in row['dept_name'])
                        title_match = (p.title or '') == row['pos_name']
                        if dept_match and title_match:
                            position = p
                            break

                if position:
                    position.apply_count = row['apply_count']
                    if position.recruitment_count and position.recruitment_count > 0:
                        position.competition_ratio = round(
                            row['apply_count'] / position.recruitment_count, 1
                        )
                    matched_ids.add(position.id)
                    matched += 1
                else:
                    unmatched += 1
                    if unmatched <= 2:
                        print(f'    未匹配: {row["dept_name"]}[{row["dept_code"]}] - {row["pos_name"]}[{row["pos_code"]}]')

            await db.flush()
            total_matched += matched
            total_unmatched += unmatched
            pct = matched / (matched + unmatched) * 100 if (matched + unmatched) > 0 else 0
            print(f'  {city:5s}: 匹配 {matched:4d}, 未匹配 {unmatched:3d} ({pct:.0f}%)')

        await db.commit()
        print(f'\n总计: 匹配 {total_matched}, 未匹配 {total_unmatched}')


async def main():
    fp = '/Users/chaim/Desktop/江苏事业单位2023~2025年进面分数线/2025年/[报名数据]2025年江苏事业单位统考报名人数汇总.xlsx'
    wb = openpyxl.load_workbook(fp, data_only=True)

    print('=== 解析报名数据 ===')
    apply_data = parse_all_sheets(wb)
    wb.close()

    print('\n=== 合并到数据库 ===')
    await merge_apply_count(apply_data)


if __name__ == '__main__':
    asyncio.run(main())
