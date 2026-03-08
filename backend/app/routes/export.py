"""数据导出路由 - 支持 Excel 导出学员/财务/考勤"""
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, extract
from datetime import date
from typing import Optional
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.database import get_db
from app.models.student import Student
from app.models.finance import FinanceRecord
from app.models.attendance import Attendance
from app.models.user import User
from app.middleware.auth import get_current_user

router = APIRouter(prefix="/api/v1/export", tags=["数据导出"])

# 导出数据量上限
MAX_EXPORT_ROWS = 10000

# 通用样式
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, col_count: int):
    """给表头行添加样式"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _to_streaming(wb: Workbook, filename: str) -> StreamingResponse:
    """将 Workbook 转为流式响应"""
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/students")
async def export_students(
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """导出学员列表 Excel"""
    query = select(Student).where(Student.deleted_at.is_(None))
    if status:
        query = query.where(Student.status == status)
    query = query.order_by(Student.id.desc()).limit(MAX_EXPORT_ROWS)

    result = await db.execute(query)
    students = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "学员列表"

    headers = ["ID", "姓名", "手机号", "微信", "性别", "学历", "专业",
               "报考类型", "班次", "入学日期", "状态", "缴费状态",
               "实际金额", "督学老师", "最近联系", "备注"]
    ws.append(headers)
    _style_header(ws, len(headers))

    status_map = {"active": "在读", "inactive": "暂停", "graduated": "结业",
                  "lead": "线索", "trial": "试听", "dropped": "退出"}

    for s in students:
        ws.append([
            s.id, s.name, s.phone, s.wechat, s.gender, s.education, s.major,
            s.exam_type, s.class_name,
            s.enrollment_date.isoformat() if s.enrollment_date else "",
            status_map.get(s.status, s.status),
            s.payment_status or "",
            float(s.actual_price) if s.actual_price else "",
            "",  # supervisor name would need join
            s.last_contact_date.isoformat() if s.last_contact_date else "",
            s.notes or "",
        ])

    # 设置列宽
    widths = [6, 10, 14, 14, 6, 8, 15, 12, 10, 12, 8, 8, 10, 10, 12, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w

    return _to_streaming(wb, "students.xlsx")


@router.get("/finance")
async def export_finance(
    record_type: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """导出财务记录 Excel"""
    query = select(FinanceRecord).where(FinanceRecord.deleted_at.is_(None))
    if record_type:
        query = query.where(FinanceRecord.record_type == record_type)
    if year:
        query = query.where(extract("year", FinanceRecord.record_date) == year)
    if month:
        query = query.where(extract("month", FinanceRecord.record_date) == month)
    query = query.order_by(FinanceRecord.record_date.desc()).limit(MAX_EXPORT_ROWS)

    result = await db.execute(query)
    records = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "财务记录"

    headers = ["ID", "类型", "分类", "金额", "日期", "说明", "支付方式", "收据号"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in records:
        ws.append([
            r.id,
            "收入" if r.record_type == "income" else "支出",
            r.category,
            float(r.amount),
            r.record_date.isoformat(),
            r.description or "",
            r.payment_method or "",
            r.receipt_no or "",
        ])

    # 金额列格式
    for row in range(2, len(records) + 2):
        ws.cell(row=row, column=4).number_format = "#,##0.00"

    widths = [6, 8, 12, 12, 12, 25, 10, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    return _to_streaming(wb, "finance.xlsx")


@router.get("/attendances")
async def export_attendances(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """导出考勤记录 Excel"""
    query = (
        select(Attendance, Student.name)
        .join(Student, Attendance.student_id == Student.id)
        .where(Attendance.deleted_at.is_(None))
    )
    if start_date:
        query = query.where(Attendance.attendance_date >= start_date)
    if end_date:
        query = query.where(Attendance.attendance_date <= end_date)
    query = query.order_by(Attendance.attendance_date.desc()).limit(MAX_EXPORT_ROWS)

    result = await db.execute(query)
    rows = result.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "考勤记录"

    headers = ["ID", "学员姓名", "日期", "状态", "备注"]
    ws.append(headers)
    _style_header(ws, len(headers))

    status_map = {"present": "出勤", "absent": "缺勤", "late": "迟到", "leave": "请假"}

    for att, student_name in rows:
        ws.append([
            att.id,
            student_name,
            att.attendance_date.isoformat(),
            status_map.get(att.status, att.status),
            att.notes or "",
        ])

    widths = [6, 12, 12, 8, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    return _to_streaming(wb, "attendances.xlsx")
