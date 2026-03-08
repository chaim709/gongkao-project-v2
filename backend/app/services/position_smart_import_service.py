"""
智能岗位导入服务 - 支持多文件自动识别、列名标准化、智能合并
"""
from io import BytesIO
from typing import Dict, List, Any, Optional, Tuple
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from app.models.position import Position
import re


class PositionSmartImportService:
    """智能岗位导入服务"""

    # ===== 列名标准化映射 =====
    COLUMN_ALIASES = {
        # 地市
        '地市': 'city', '地': 'city', '地区': 'city',
        # 隶属关系
        '隶属关系': 'affiliation', '隶属  关系': 'affiliation',
        # 地区代码
        '地区代码': 'district_code', '地区  代码': 'district_code',
        # 地区名称（映射到 location）
        '地区名称': 'location',
        # 单位代码
        '单位代码': 'department_code', '部门代码': 'department_code',
        # 单位名称
        '单位名称': 'department', '部门名称': 'department',
        '招考部门': 'department', '招录机关': 'department',
        # 职位代码
        '职位代码': 'position_code',
        # 职位名称
        '职位名称': 'title',
        # 职位简介
        '职位简介': 'description', '职位描述': 'description',
        # 考试类别
        '职位类别': 'exam_category', '考试类别': 'exam_category',
        '职位类型': 'exam_category', '岗位类别': 'exam_category',
        # 开考比例
        '开考比例': 'open_ratio',
        # 招录人数
        '招考人数': 'recruitment_count', '招录人数': 'recruitment_count',
        '录用计划数': 'recruitment_count',
        # 学历
        '学历': 'education', '学\u3000历': 'education', '学历要求': 'education',
        # 专业
        '专业': 'major', '专\u3000业': 'major', '专业要求': 'major',
        # 其他条件
        '其它': 'other_requirements', '其\u3000它': 'other_requirements',
        '其他条件': 'other_requirements', '其他要求': 'other_requirements',
        '备注': 'other_requirements',
        # 单位性质
        '单位性质': 'department_type', '地市+系统': 'department_type',
        # 报名人数
        '报名成功人数': 'successful_applicants',
        '审核通过人数': 'successful_applicants',
        '报名人数': 'successful_applicants',
        '报名人数2': 'successful_applicants',
        '合格人数': 'successful_applicants',
        # 竞争比
        '竞争比': 'competition_ratio',
        # 进面分数线
        '进面最低分': 'min_interview_score', '最低进面分': 'min_interview_score',
        '进面分': 'min_interview_score',
        '进面最高分': 'max_interview_score', '最高进面分': 'max_interview_score',
        # 行测/申论
        '最高行测': 'max_xingce_score',
        '最高申论': 'max_shenlun_score',
        # 专业技能
        '专业/技能': 'professional_skills', '专业技能': 'professional_skills',
        # 额外字段（可能出现但不需要的）
        '年份': '_year', '省份': '_province',
        '考试类型': '_exam_type', '系统分布': '_system',
        '用人司局': 'hiring_unit', '区县': '_county',
        # 国考扩展字段
        '机构层级': 'institution_level', '职位属性': 'position_attribute',
        '职位分布': 'position_distribution', '面试人员比例': 'interview_ratio',
        '落户地点': 'settlement_location', '服务基层项目工作经历': 'grassroots_project',
        # 所在地（部分职位表独有）
        '所在地': 'location', '职位所属地区': 'location',
        '地区  名称': 'location',
    }

    # 文件类型特征列
    SCORE_INDICATORS = {'进面最低分', '最低进面分', '进面分', '进面最高分', '最高进面分'}
    APPLY_INDICATORS = {'报名成功人数', '审核通过人数', '报名人数', '报名人数2', '合格人数'}

    # ===== 文件类型检测 =====
    @classmethod
    def detect_file_type(cls, headers: List[str]) -> str:
        """
        检测文件类型:
        - position: 职位表（只有基础信息）
        - application: 报名人数表（基础+报名数据）
        - score: 进面分数线表（分数线数据）
        - complete: 完整数据（全部字段）
        """
        header_set = set(h.strip() if h else '' for h in headers)

        has_score = bool(header_set & cls.SCORE_INDICATORS)
        has_apply = bool(header_set & cls.APPLY_INDICATORS)
        has_position_code = '职位代码' in header_set
        has_recruit = bool(header_set & {'招考人数', '招录人数', '录用计划数'})

        if has_score and has_apply and has_recruit:
            return 'complete'
        if has_score and not has_recruit:
            return 'score'
        if has_apply:
            return 'application'
        if has_position_code or has_recruit:
            return 'position'

        return 'unknown'

    @classmethod
    def detect_header_row(cls, ws) -> Tuple[int, List[str]]:
        """检测实际表头行（可能第1行是广告）"""
        row1 = [str(c.value).strip() if c.value else '' for c in ws[1]]
        row2_cells = list(ws[2]) if ws.max_row >= 2 else []
        row2 = [str(c.value).strip() if c.value else '' for c in row2_cells]

        # 检查 row1 是否像表头（包含已知列名）
        known_cols = set(cls.COLUMN_ALIASES.keys())
        row1_matches = sum(1 for h in row1 if h in known_cols)
        row2_matches = sum(1 for h in row2 if h in known_cols)

        if row1_matches >= 3:
            return 1, row1
        if row2_matches >= 3:
            return 2, row2

        # 如果都不明显，默认第1行
        return 1, row1

    # ===== 列映射 =====
    @classmethod
    def map_columns(cls, headers: List[str]) -> Dict[int, str]:
        """将 Excel 列索引映射到模型字段名"""
        mapping = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            clean = header.strip().replace('\n', '').replace('\r', '')
            if clean in cls.COLUMN_ALIASES:
                field = cls.COLUMN_ALIASES[clean]
                mapping[idx] = field
        return mapping

    # ===== 解析文件 =====
    @classmethod
    def parse_file(cls, content: bytes, sheet_name: str = None) -> Dict[str, Any]:
        """
        解析单个 Excel 文件，返回标准化数据。
        返回: {type, header_row, column_mapping, rows: [{field: value}]}
        """
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        header_row_num, headers = cls.detect_header_row(ws)
        file_type = cls.detect_file_type(headers)
        column_mapping = cls.map_columns(headers)

        rows = []
        str_fields = {'district_code', 'department_code', 'position_code'}

        for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
            if not any(row):
                continue

            record = {}
            for col_idx, field in column_mapping.items():
                if col_idx < len(row):
                    value = row[col_idx]
                    if value is None:
                        continue
                    if field.startswith('_'):
                        record[field] = value
                        continue
                    if field in str_fields:
                        record[field] = str(value).strip()
                    elif field in ('recruitment_count', 'successful_applicants'):
                        try:
                            record[field] = int(float(value))
                        except (ValueError, TypeError):
                            pass
                    elif field in ('competition_ratio', 'min_interview_score',
                                   'max_interview_score', 'max_xingce_score',
                                   'max_shenlun_score'):
                        try:
                            record[field] = float(value)
                        except (ValueError, TypeError):
                            pass
                    else:
                        record[field] = str(value).strip() if value is not None else None

            if record.get('title') or record.get('department') or record.get('position_code'):
                rows.append(record)

        wb.close()
        return {
            'type': file_type,
            'header_row': header_row_num,
            'column_count': len(column_mapping),
            'row_count': len(rows),
            'columns_detected': {v: k for k, v in column_mapping.items() if not v.startswith('_')},
            'rows': rows,
        }

    # ===== 进面分数线合并 =====
    @classmethod
    def merge_scores(
        cls,
        position_rows: List[Dict],
        score_rows: List[Dict],
    ) -> Dict[str, Any]:
        """
        将进面分数线数据合并到职位数据中。
        使用多策略匹配：城市+职位代码 → 地区名/部门名缩小 → 模糊匹配
        """
        # 建立职位索引
        # key: (city, position_code) -> [position_row_indices]
        city_code_index: Dict[Tuple[str, str], List[int]] = {}
        for idx, pos in enumerate(position_rows):
            city = str(pos.get('city', '')).strip()
            code = str(pos.get('position_code', '')).strip()
            if city and code:
                key = (city, code)
                city_code_index.setdefault(key, []).append(idx)

        matched = 0
        unmatched = []

        for score in score_rows:
            score_city = str(score.get('city', '')).strip()
            score_code = str(score.get('position_code', '')).strip()
            score_dept = str(score.get('department', score.get('_county', ''))).strip()
            score_title = str(score.get('title', '')).strip()

            if not score_code:
                unmatched.append({'reason': '无职位代码', 'data': score})
                continue

            # 策略1: 城市 + 职位代码
            candidates_indices = city_code_index.get((score_city, score_code), [])

            # 城市名可能少个"市"字
            if not candidates_indices:
                candidates_indices = city_code_index.get((score_city + '市', score_code), [])

            if len(candidates_indices) == 1:
                # 唯一匹配
                cls._apply_score(position_rows[candidates_indices[0]], score)
                matched += 1
                continue

            if len(candidates_indices) > 1:
                # 多个候选，用部门名/地区名缩小
                match_idx = cls._narrow_by_name(
                    position_rows, candidates_indices, score_dept, score_title
                )
                if match_idx is not None:
                    cls._apply_score(position_rows[match_idx], score)
                    matched += 1
                    continue

            # 策略2: 全局搜索同代码的，用部门名匹配
            if not candidates_indices:
                all_with_code = []
                for key, indices in city_code_index.items():
                    if key[1] == score_code:
                        all_with_code.extend(indices)

                if all_with_code:
                    match_idx = cls._narrow_by_name(
                        position_rows, all_with_code, score_dept, score_title
                    )
                    if match_idx is not None:
                        cls._apply_score(position_rows[match_idx], score)
                        matched += 1
                        continue

            unmatched.append({
                'reason': f'无法匹配 city={score_city} code={score_code} dept={score_dept}',
                'data': {
                    'city': score_city,
                    'department': score_dept,
                    'title': score_title,
                    'position_code': score_code,
                    'min_interview_score': score.get('min_interview_score'),
                    'max_interview_score': score.get('max_interview_score'),
                }
            })

        return {
            'total': len(score_rows),
            'matched': matched,
            'unmatched_count': len(unmatched),
            'unmatched_details': unmatched[:20],
        }

    @classmethod
    def _narrow_by_name(
        cls,
        positions: List[Dict],
        indices: List[int],
        score_dept: str,
        score_title: str,
    ) -> Optional[int]:
        """在多个候选中通过部门名/职位名缩小到唯一"""
        if not score_dept and not score_title:
            return indices[0] if len(indices) == 1 else None

        # 解析进面表的部门名: "丹阳市-丹北镇政府" -> "丹北镇政府"
        dept_parts = score_dept.split('-') if '-' in score_dept else [score_dept]
        dept_area = dept_parts[0].strip() if len(dept_parts) > 1 else ''
        dept_unit = dept_parts[-1].strip()

        # 按 地区名/location 缩小
        if dept_area:
            area_filtered = [
                i for i in indices
                if dept_area in str(positions[i].get('location', ''))
                or str(positions[i].get('location', '')) in dept_area
            ]
            if area_filtered:
                indices = area_filtered

        # 按 部门名 缩小
        if dept_unit:
            name_filtered = [
                i for i in indices
                if dept_unit in str(positions[i].get('department', ''))
                or str(positions[i].get('department', '')) in dept_unit
            ]
            if len(name_filtered) >= 1:
                indices = name_filtered

        # 按 职位名 缩小
        if score_title and len(indices) > 1:
            title_filtered = [
                i for i in indices
                if score_title == str(positions[i].get('title', ''))
                or score_title in str(positions[i].get('title', ''))
                or str(positions[i].get('title', '')) in score_title
            ]
            if title_filtered:
                indices = title_filtered

        return indices[0] if indices else None

    @staticmethod
    def _apply_score(position: Dict, score: Dict):
        """将分数线数据合并到职位记录"""
        for field in ('min_interview_score', 'max_interview_score',
                      'max_xingce_score', 'max_shenlun_score'):
            if score.get(field) is not None:
                position[field] = score[field]

    # ===== 主入口：智能导入 =====
    @classmethod
    async def smart_import(
        cls,
        db: AsyncSession,
        files: List[Tuple[str, bytes]],
        year: int,
        exam_type: str,
    ) -> Dict[str, Any]:
        """
        智能导入入口。

        Args:
            files: [(filename, content), ...]
            year: 年份
            exam_type: 考试类型（省考/事业单位/国考）
        """
        # Step 1: 解析所有文件
        parsed_files = []
        for filename, content in files:
            parsed = cls.parse_file(content)
            parsed['filename'] = filename
            parsed_files.append(parsed)

        # Step 2: 分类文件
        position_data = []
        application_data = []
        score_data = []

        for pf in parsed_files:
            if pf['type'] == 'complete':
                position_data.extend(pf['rows'])
            elif pf['type'] == 'position':
                position_data.extend(pf['rows'])
            elif pf['type'] == 'application':
                application_data.extend(pf['rows'])
            elif pf['type'] == 'score':
                score_data.extend(pf['rows'])

        # Step 3: 合并报名人数（结构相同，直接覆盖）
        if application_data and not position_data:
            # 如果只有报名人数表，当作位置数据用
            position_data = application_data
        elif application_data and position_data:
            # 两者都有，将报名数据合并到职位数据
            app_index = {}
            for app in application_data:
                key = (
                    str(app.get('district_code', '')),
                    str(app.get('department_code', '')),
                    str(app.get('position_code', '')),
                )
                app_index[key] = app

            for pos in position_data:
                key = (
                    str(pos.get('district_code', '')),
                    str(pos.get('department_code', '')),
                    str(pos.get('position_code', '')),
                )
                if key in app_index:
                    app = app_index[key]
                    if app.get('successful_applicants') is not None:
                        pos['successful_applicants'] = app['successful_applicants']
                    if app.get('competition_ratio') is not None:
                        pos['competition_ratio'] = app['competition_ratio']

        # Step 4: 合并进面分数线
        merge_result = None
        if score_data and position_data:
            merge_result = cls.merge_scores(position_data, score_data)

        # Step 5: 入库（upsert）
        inserted = 0
        updated = 0
        errors = []

        # 有效的 Position 模型字段
        valid_fields = {c.name for c in Position.__table__.columns if c.name != 'id'}

        for idx, row in enumerate(position_data):
            try:
                # 设置年份和考试类型
                row['year'] = row.pop('_year', None) or year
                if isinstance(row['year'], str):
                    row['year'] = int(row['year'])
                row['exam_type'] = row.pop('_exam_type', None) or exam_type

                # 自动计算竞争比（公式值可能丢失）
                if not row.get('competition_ratio'):
                    applicants = row.get('successful_applicants')
                    recruit = row.get('recruitment_count')
                    if applicants and recruit and recruit > 0:
                        row['competition_ratio'] = round(applicants / recruit, 1)

                # 清理内部字段
                clean_row = {k: v for k, v in row.items()
                             if not k.startswith('_') and k in valid_fields}

                # 尝试查找已存在的记录（用复合键）
                dc = clean_row.get('district_code', '')
                dept_c = clean_row.get('department_code', '')
                pc = clean_row.get('position_code', '')

                if dc and dept_c and pc:
                    existing = (await db.execute(
                        select(Position).where(and_(
                            Position.year == clean_row['year'],
                            Position.exam_type == clean_row['exam_type'],
                            Position.district_code == dc,
                            Position.department_code == dept_c,
                            Position.position_code == pc,
                        ))
                    )).scalar_one_or_none()

                    if existing:
                        # 更新非空字段
                        for field, value in clean_row.items():
                            if value is not None and field not in ('year', 'exam_type',
                                                                     'district_code',
                                                                     'department_code',
                                                                     'position_code'):
                                setattr(existing, field, value)
                        updated += 1
                        continue

                # 新记录
                position = Position(**clean_row)
                db.add(position)
                inserted += 1

            except Exception as e:
                errors.append(f"第{idx + 1}条: {str(e)}")

        await db.commit()

        return {
            'detected_files': [
                {
                    'filename': pf['filename'],
                    'type': pf['type'],
                    'type_label': {
                        'position': '职位表',
                        'application': '报名人数表',
                        'score': '进面分数线表',
                        'complete': '完整数据',
                        'unknown': '未知类型',
                    }.get(pf['type'], '未知'),
                    'rows': pf['row_count'],
                    'columns': pf['column_count'],
                }
                for pf in parsed_files
            ],
            'import_result': {
                'total': len(position_data),
                'inserted': inserted,
                'updated': updated,
                'errors': errors[:20],
            },
            'merge_result': merge_result,
        }
