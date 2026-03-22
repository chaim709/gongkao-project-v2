"""事业编数据导入服务 - 解析江苏事业单位统考岗位表"""
from dataclasses import dataclass
from typing import Dict, Any, Set
import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.models.position import Position


# 事业编旧版合并表：三行表头，数据从第 4 行开始
LEGACY_SHIYE_COLUMN_MAP = {
    0: 'city',               # 地市
    1: 'location',           # 区县
    2: 'supervising_dept',   # 主管部门
    3: 'department',         # 招聘单位 名称
    4: 'department_code',    # 招聘单位 单位代码
    5: 'funding_source',     # 招聘单位 经费来源
    6: 'title',              # 招聘岗位 岗位名称
    7: 'position_code',      # 招聘岗位 岗位代码
    8: 'exam_category',      # 招聘岗位 笔试类别
    9: 'description',        # 招聘岗位 岗位描述
    10: 'recruitment_count', # 招聘人数
    11: 'exam_ratio',        # 开考比例
    12: 'recruitment_target',# 招聘对象
    13: 'education',         # 招聘条件 学历
    14: 'major',             # 招聘条件 专业
    15: 'other_requirements',# 招聘条件 其他条件
    16: 'apply_count',       # 报名人数
    17: '_competition_ratio',# 竞争比（需解析）
    18: 'min_interview_score',# 进面最低分
    19: 'max_interview_score',# 进面最高分
}

# 事业编总表：单行表头，数据从第 2 行开始
TOTAL_TABLE_COLUMN_MAP = {
    0: 'city',
    1: 'location',
    2: 'supervising_dept',
    3: 'department',
    4: 'position_code',
    5: 'title',
    6: 'description',
    7: 'exam_category',
    8: 'position_level',
    9: 'funding_source',
    10: 'recruitment_count',
    11: 'education',
    12: 'degree',
    13: 'major',
    14: 'other_requirements',
    15: 'recruitment_target',
    16: 'exam_ratio',
    17: 'exam_weight_ratio',
    18: 'interview_ratio',
    19: 'remark',
    20: 'apply_count',
    21: '_competition_ratio',
    22: 'min_interview_score',
    23: 'max_interview_score',
}

# 需要向下填充的列（合并单元格导致后续行为空）
# 注意：旧版 Col 4 (department_code) 不做填充，每个岗位有独立的单位代码
LEGACY_FILL_DOWN_COLS = {0, 1, 2, 3, 5}
TOTAL_TABLE_FILL_DOWN_COLS = {0, 1, 2, 3, 9}

# 城市名标准化
CITY_NORMALIZE = {
    '省属': '省属', '南京': '南京市', '苏州': '苏州市', '无锡': '无锡市',
    '常州': '常州市', '南通': '南通市', '徐州': '徐州市', '扬州': '扬州市',
    '镇江': '镇江市', '盐城': '盐城市', '泰州': '泰州市', '淮安': '淮安市',
    '连云港': '连云港市', '宿迁': '宿迁市',
}


@dataclass(frozen=True)
class ShiyeImportLayout:
    name: str
    data_start_row: int
    column_map: dict[int, str]
    fill_down_cols: set[int]


LEGACY_LAYOUT = ShiyeImportLayout(
    name="legacy_merged_sheet",
    data_start_row=4,
    column_map=LEGACY_SHIYE_COLUMN_MAP,
    fill_down_cols=LEGACY_FILL_DOWN_COLS,
)

TOTAL_TABLE_LAYOUT = ShiyeImportLayout(
    name="normalized_total_table",
    data_start_row=2,
    column_map=TOTAL_TABLE_COLUMN_MAP,
    fill_down_cols=TOTAL_TABLE_FILL_DOWN_COLS,
)


class ShiyeImportService:
    """事业编 .xlsx 文件导入服务"""

    @staticmethod
    def _clean_header(value: Any) -> str:
        if value is None:
            return ""
        return str(value).replace("\n", "").replace("\r", "").strip()

    @classmethod
    def detect_layout(cls, ws) -> ShiyeImportLayout:
        row1 = [cls._clean_header(cell.value) for cell in ws[1]]
        row2 = [cls._clean_header(cell.value) for cell in ws[2]] if ws.max_row >= 2 else []
        row3 = [cls._clean_header(cell.value) for cell in ws[3]] if ws.max_row >= 3 else []

        row1_set = {value for value in row1 if value}
        row2_set = {value for value in row2 if value}
        row3_set = {value for value in row3 if value}

        if {
            "地市",
            "区县",
            "主管部门",
            "招录单位",
            "岗位名称",
            "经费来源",
            "招聘对象",
            "报名人数",
        }.issubset(row1_set):
            return TOTAL_TABLE_LAYOUT

        if (
            {"地市", "区县", "主管部门", "招聘单位", "招聘岗位", "招聘对象", "报名人数"}
            .issubset(row2_set)
            and {"名称", "单位代码", "经费来源", "岗位名称", "岗位代码", "笔试类别", "学历", "专业"}
            .issubset(row3_set)
        ):
            return LEGACY_LAYOUT

        raise ValueError("无法识别事业编导入文件格式")

    @staticmethod
    def _parse_row(filled_row: list[Any], column_map: dict[int, str]) -> Dict[str, Any]:
        row_data: Dict[str, Any] = {}

        for ci, field in column_map.items():
            if ci >= len(filled_row):
                continue
            val = filled_row[ci]

            if field == 'recruitment_count':
                try:
                    row_data[field] = int(val) if val and str(val).strip() not in ('——', '-', '') else 1
                except (ValueError, TypeError):
                    row_data[field] = 1
            elif field == 'apply_count':
                try:
                    row_data[field] = int(val) if val and str(val).strip() not in ('——', '-', '') else None
                except (ValueError, TypeError):
                    row_data[field] = None
            elif field in ('min_interview_score', 'max_interview_score'):
                try:
                    if val and str(val).strip() not in ('——', '-', '', '未公示', '/'):
                        row_data[field] = float(val)
                    else:
                        row_data[field] = None
                except (ValueError, TypeError):
                    row_data[field] = None
            elif field == '_competition_ratio':
                try:
                    if isinstance(val, (int, float)):
                        row_data['competition_ratio'] = float(val)
                    elif val and str(val).strip() not in ('——', '-', '', '#VALUE!'):
                        s = str(val).strip()
                        if ':' in s:
                            row_data['competition_ratio'] = float(s.split(':')[0])
                        else:
                            row_data['competition_ratio'] = float(s)
                except (ValueError, TypeError):
                    pass
            else:
                row_data[field] = str(val).strip() if val is not None else None

        return row_data

    @classmethod
    async def import_file(
        cls, db: AsyncSession, file_path: str, year: int
    ) -> Dict[str, Any]:
        """导入事业编岗位表（含竞争比和进面分）"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        layout = cls.detect_layout(ws)

        inserted = 0
        updated = 0
        skipped = 0

        # 检查是否已有该年份的事业编数据（决定是否需要 upsert）
        existing_count = (await db.execute(
            select(Position.id).where(
                Position.year == year, Position.exam_type == '事业单位'
            ).limit(1)
        )).scalar_one_or_none()
        need_upsert = existing_count is not None

        # 向下填充缓存：记住上一行的值，用于填充合并单元格
        prev_values: Dict[int, Any] = {}
        # 追踪已匹配的 DB 记录 ID，避免同一条被多次更新
        matched_ids: Set[int] = set()

        for ri, row in enumerate(
            ws.iter_rows(min_row=layout.data_start_row, values_only=True),
            start=layout.data_start_row,
        ):
            try:
                # 处理向下填充：合并单元格只有首行有值
                filled_row = list(row)
                for ci in layout.fill_down_cols:
                    if ci < len(filled_row):
                        if filled_row[ci] is not None and str(filled_row[ci]).strip() != '':
                            prev_values[ci] = filled_row[ci]
                        else:
                            filled_row[ci] = prev_values.get(ci)

                row_data = cls._parse_row(filled_row, layout.column_map)

                # 城市名标准化
                if row_data.get('city'):
                    raw_city = row_data['city'].replace('市', '')
                    row_data['city'] = CITY_NORMALIZE.get(raw_city, row_data['city'])

                # 固定字段
                row_data['year'] = year
                row_data['exam_type'] = '事业单位'
                row_data['province'] = '江苏省'

                # 跳过无效行（title 是每行独有的，不应被填充后仍为空）
                if not row_data.get('title'):
                    if row_data.get('description'):
                        row_data['title'] = row_data['description'][:100]
                    elif row_data.get('position_code'):
                        row_data['title'] = f"岗位{row_data['position_code']}"
                    else:
                        skipped += 1
                        continue

                # department 经过填充后应该有值
                if not row_data.get('department'):
                    skipped += 1
                    continue

                # Upsert 查找（仅在重复导入时需要）
                existing = None
                if need_upsert:
                    city_val = row_data.get('city')
                    dept_code = row_data.get('department_code')
                    pos_code = row_data.get('position_code')

                    if dept_code and pos_code:
                        results = (await db.execute(
                            select(Position).where(
                                Position.year == year,
                                Position.exam_type == '事业单位',
                                Position.city == city_val,
                                Position.department_code == dept_code,
                                Position.position_code == pos_code,
                            )
                        )).scalars().all()
                        for r in results:
                            if r.id not in matched_ids:
                                existing = r
                                break
                    else:
                        conditions = [
                            Position.year == year,
                            Position.exam_type == '事业单位',
                            Position.city == city_val,
                            Position.department == row_data.get('department'),
                            Position.title == row_data.get('title'),
                        ]
                        if pos_code:
                            conditions.append(Position.position_code == pos_code)
                        results = (await db.execute(
                            select(Position).where(*conditions)
                        )).scalars().all()
                        for r in results:
                            if r.id not in matched_ids:
                                existing = r
                                break

                if existing:
                    matched_ids.add(existing.id)
                    for key, val in row_data.items():
                        if val is not None:
                            setattr(existing, key, val)
                    updated += 1
                else:
                    db.add(Position(**row_data))
                    inserted += 1

                if (inserted + updated) % 500 == 0:
                    await db.flush()

            except Exception as e:
                skipped += 1
                if skipped <= 5:
                    print(f"  跳过行 {ri}: {e}")

        await db.commit()
        wb.close()

        print(f"  插入: {inserted}, 更新: {updated}, 跳过: {skipped}")
        return {
            'year': year,
            'inserted': inserted,
            'updated': updated,
            'skipped': skipped,
        }
