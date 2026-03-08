"""岗位导入服务"""
from io import BytesIO
from typing import List, Dict, Any
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from app.models.position import Position
from app.exceptions.business import BusinessError


class PositionImportService:
    """岗位导入服务"""

    @staticmethod
    def normalize_education(education: str) -> str:
        """标准化学历为四个标准选项"""
        if not education:
            return education

        edu = str(education).strip()

        if any(k in edu for k in ['研究生', '硕士', '博士']):
            return '研究生及以上'

        if any(k in edu for k in ['本科', '学士']):
            return '本科及以上'

        if any(k in edu for k in ['大专', '专科', '高职']):
            return '大专及以上'

        if any(k in edu for k in ['高中', '中专', '中职']):
            return '高中（中专）'

        return education

    # Excel 列定义
    COLUMNS = [
        ("year", "年份"),
        ("exam_type", "考试类型"),
        ("city", "地市"),
        ("affiliation", "隶属关系"),
        ("district_code", "地区代码"),
        ("department", "单位名称"),
        ("department_code", "单位代码"),
        ("department_type", "单位性质"),
        ("position_code", "职位代码"),
        ("title", "职位名称"),
        ("description", "职位简介"),
        ("exam_category", "职位类别"),
        ("recruitment_count", "招录人数"),
        ("education", "学历要求"),
        ("major", "专业要求"),
        ("degree", "学位要求"),
        ("political_status", "政治面貌"),
        ("work_experience", "工作经验"),
        ("other_requirements", "其他要求"),
        ("location", "工作地点"),
        ("successful_applicants", "成功报名人数"),
        ("competition_ratio", "竞争比"),
        ("min_interview_score", "最低进面分数线"),
        ("max_interview_score", "最高进面分数线"),
        ("max_xingce_score", "最高行测"),
        ("max_shenlun_score", "最高申论"),
        ("professional_skills", "专业技能"),
    ]

    @classmethod
    def generate_template(cls) -> BytesIO:
        """生成导入模板"""
        wb = Workbook()
        ws = wb.active
        ws.title = "岗位导入模板"

        # 写入表头
        for idx, (_, label) in enumerate(cls.COLUMNS, start=1):
            ws.cell(row=1, column=idx, value=label)

        # 写入示例数据
        example = [
            2026, "省考", "南京市", "省直", "000100", "省委办公厅", "502", "党群机关",
            "01", "一级科员", "从事财务工作", "A", 1, "本科及以上", "财会类", "学士",
            "不限", "无要求", "取得相应学位", "南京市",
            150, 75.5, 125.3, 138.7, 78.5, 72.3, "无"
        ]
        for idx, value in enumerate(example, start=1):
            ws.cell(row=2, column=idx, value=value)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @classmethod
    async def import_positions(cls, db: AsyncSession, file_content: bytes) -> Dict[str, Any]:
        """导入岗位数据"""
        wb = load_workbook(BytesIO(file_content))
        ws = wb.active

        # 验证表头
        headers = [cell.value for cell in ws[1]]
        expected_headers = [label for _, label in cls.COLUMNS]
        if headers[:len(expected_headers)] != expected_headers:
            raise BusinessError(5001, "Excel 表头格式不正确，请使用标准模板")

        # 解析数据
        positions = []
        errors = []

        # 需要转换为字符串的字段
        str_fields = {'district_code', 'department_code', 'position_code'}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # 跳过空行
                continue

            try:
                position_data = {}
                for idx, (field, _) in enumerate(cls.COLUMNS):
                    value = row[idx] if idx < len(row) else None
                    if value is not None:
                        if field == 'education':
                            position_data[field] = cls.normalize_education(str(value))
                        elif field in str_fields:
                            position_data[field] = str(value)
                        else:
                            position_data[field] = value

                # 验证必填字段
                if not position_data.get("title"):
                    errors.append(f"第{row_idx}行：职位名称不能为空")
                    continue

                positions.append(Position(**position_data))
            except Exception as e:
                errors.append(f"第{row_idx}行：{str(e)}")

        # 批量插入
        if positions:
            db.add_all(positions)
            await db.commit()

        return {
            "success_count": len(positions),
            "error_count": len(errors),
            "errors": errors[:10]  # 只返回前10条错误
        }
