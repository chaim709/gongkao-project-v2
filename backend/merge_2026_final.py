"""整合2026年江苏事业单位最终版，以完整版为基准补充字段"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def clean(val):
    if val is None or val == '':
        return None
    s = str(val).strip().replace('\n', ' ')
    return s if s not in ('', '——', '-', '/') else None

def main():
    fp = '/Users/chaim/Desktop/事业单位岗位表/江苏省2026年事业单位招聘岗位总表_最终版.xlsx'
    wb_in = openpyxl.load_workbook(fp, data_only=True, read_only=True)
    ws_in = wb_in[wb_in.sheetnames[0]]

    # 读取数据
    data = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        if not row[3]:  # 招录单位为空跳过
            continue

        # 构建备注
        remark_parts = []
        if clean(row[18]):  # 是否进编
            remark_parts.append(f"编制：{clean(row[18])}")
        if clean(row[12]):  # 考察体检比例
            remark_parts.append(f"体检比例：{clean(row[12])}")
        if clean(row[26]):  # 技能测试
            remark_parts.append(f"技能测试：{clean(row[26])}")
        if clean(row[27]):  # 成绩计算
            remark_parts.append(f"成绩计算：{clean(row[27])}")
        if clean(row[28]):  # 专业加试
            remark_parts.append(f"专业加试：{clean(row[28])}")
        if clean(row[19]):  # 咨询电话
            remark_parts.append(f"咨询：{clean(row[19])}")
        if clean(row[29]):  # 单位地址
            remark_parts.append(f"地址：{clean(row[29])}")
        if clean(row[24]):  # 其他说明
            remark_parts.append(clean(row[24]))

        # 笔试科目合并
        exam_cat = clean(row[22]) or clean(row[30]) or clean(row[8])

        item = {
            'city': clean(row[0]),
            'location': clean(row[1]),
            'supervising_dept': clean(row[2]),
            'department': clean(row[3]),
            'position_code': clean(row[7]),
            'title': clean(row[6]),
            'description': clean(row[9]),
            'exam_category': exam_cat,
            'position_level': clean(row[23]),
            'funding_source': clean(row[5]),
            'recruitment_count': clean(row[10]),
            'education': clean(row[13]),
            'major': clean(row[14]),
            'recruitment_target': clean(row[15]),
            'other_requirements': clean(row[16]),
            'exam_ratio': clean(row[11]),
            'exam_weight_ratio': clean(row[17]),
            'remark': ' | '.join(remark_parts) if remark_parts else None,
        }

        # 退役士兵专项标注
        if item.get('location') and '退役' in item['location'] and not item.get('major'):
            item['major'] = '退役大学生士兵专项（不限专业）'

        data.append(item)

    wb_in.close()
    print(f'解析数据: {len(data)} 条')

    # 导出Excel
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = '2026年江苏事业单位'

    # 标准列定义（24列，与2023-2025保持一致）
    columns = [
        ('city', '地市', 10),
        ('location', '区县', 12),
        ('supervising_dept', '主管部门', 22),
        ('department', '招录单位', 25),
        ('position_code', '岗位代码', 10),
        ('title', '岗位名称', 18),
        ('description', '岗位说明', 25),
        ('exam_category', '笔试类别', 14),
        ('position_level', '岗位等级', 12),
        ('funding_source', '经费来源', 12),
        ('recruitment_count', '招聘人数', 8),
        ('education', '学历', 14),
        ('degree', '学位', 12),
        ('major', '专业要求', 30),
        ('other_requirements', '其他条件', 25),
        ('recruitment_target', '招聘对象', 12),
        ('exam_ratio', '开考比例', 10),
        ('exam_weight_ratio', '笔面试占比', 20),
        ('interview_ratio', '面试比例', 12),
        ('remark', '备注', 35),
        ('apply_count', '报名人数', 10),
        ('competition_ratio', '竞争比', 8),
        ('min_interview_score', '进面最低分', 10),
        ('max_interview_score', '进面最高分', 10),
    ]

    # 样式
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    cell_align = Alignment(vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    # 写表头
    for ci, (field, header, width) in enumerate(columns, 1):
        cell = ws_out.cell(row=1, column=ci, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws_out.column_dimensions[get_column_letter(ci)].width = width

    ws_out.freeze_panes = 'A2'
    ws_out.auto_filter.ref = f'A1:{get_column_letter(len(columns))}1'

    # 写数据
    for ri, item in enumerate(data, 2):
        for ci, (field, _, _) in enumerate(columns, 1):
            val = item.get(field, '')
            cell = ws_out.cell(row=ri, column=ci, value=val)
            cell.alignment = cell_align
            cell.border = border

    out = '/Users/chaim/Desktop/2026年江苏事业单位总表.xlsx'
    wb_out.save(out)
    wb_out.close()
    print(f'已导出: {out}')


if __name__ == '__main__':
    main()
