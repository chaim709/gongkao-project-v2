"""补充2023/2024事业单位缺失字段：考试形式和所占比例(exam_weight_ratio)、备注(remark)

直接从原始职位表读取，按 (city, department, position_code) 匹配更新
"""
import asyncio
import re
import openpyxl
from sqlalchemy import select
from app.database import AsyncSessionLocal
from app.models.position import Position

CITY_NORMALIZE = {
    '省属': '省属', '南京': '南京市', '南京市': '南京市',
    '苏州': '苏州市', '苏州市': '苏州市', '无锡': '无锡市', '无锡市': '无锡市',
    '常州': '常州市', '常州市': '常州市', '南通': '南通市', '南通市': '南通市',
    '徐州': '徐州市', '徐州市': '徐州市', '扬州': '扬州市', '扬州市': '扬州市',
    '镇江': '镇江市', '镇江市': '镇江市', '盐城': '盐城市', '盐城市': '盐城市',
    '泰州': '泰州市', '泰州市': '泰州市', '淮安': '淮安市', '淮安市': '淮安市',
    '连云港': '连云港市', '连云港市': '连云港市', '宿迁': '宿迁市', '宿迁市': '宿迁市',
    '海事局': '省属', '地震局': '省属',
}

INVALID = ('', '——', '-', '/', 'None')


def normalize_city(raw):
    if not raw:
        return None
    raw = str(raw).replace('\n', '').strip()
    return CITY_NORMALIZE.get(raw)


def clean(val):
    if val is None:
        return None
    s = str(val).strip().replace('\n', ' ')
    return s if s and s not in INVALID else None


def codes_equal(a, b):
    if not a or not b:
        return False
    a, b = str(a).strip(), str(b).strip()
    if a == b:
        return True
    try:
        return int(a) == int(b)
    except ValueError:
        return False


async def merge_year(year: int):
    fp = f'/Users/chaim/Desktop/江苏事业单位2023~2025年进面分数线/{year}年/[职位表]{year}年江苏事业单位统考职位表汇总.xlsx'
    wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    # COL_MAP: 0=city, 3=department, 5=position_code, 6=title
    # 新增: 16=考试形式和所占比例, 19=备注
    excel_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        city = normalize_city(row[0])
        dept = clean(row[3])
        if not city or not dept:
            continue
        code = clean(row[5])
        title = clean(row[6])
        ratio = clean(row[16])
        remark = clean(row[19])
        if ratio or remark:
            excel_rows.append({
                'city': city, 'department': dept, 'position_code': code,
                'title': title, 'exam_weight_ratio': ratio, 'remark': remark,
            })
    wb.close()
    print(f'  Excel有效行: {len(excel_rows)}')

    by_city = {}
    for r in excel_rows:
        by_city.setdefault(r['city'], []).append(r)

    async with AsyncSessionLocal() as db:
        total_matched = 0
        total_unmatched = 0

        for city in sorted(by_city.keys()):
            rows = by_city[city]
            positions = (await db.execute(
                select(Position).where(
                    Position.year == year,
                    Position.exam_type == '事业单位',
                    Position.city == city,
                )
            )).scalars().all()

            by_dept_code = {}
            for p in positions:
                dept_clean = (p.department or '').replace('\n', '').replace('\r', '')
                key = (dept_clean, p.position_code or '')
                by_dept_code.setdefault(key, []).append(p)

            matched = 0
            unmatched = 0
            matched_ids = set()

            for r in rows:
                position = None

                # 策略1: dept + code 精确
                if r['department'] and r['position_code']:
                    candidates = by_dept_code.get((r['department'], r['position_code']), [])
                    for c in candidates:
                        if c.id not in matched_ids:
                            position = c
                            break

                # 策略2: dept子串 + code
                if not position and r['position_code']:
                    for p in positions:
                        if p.id in matched_ids:
                            continue
                        p_dept = (p.department or '').replace('\n', '')
                        dept_match = (r['department'] in p_dept or p_dept in r['department'])
                        code_match = codes_equal(p.position_code, r['position_code'])
                        if dept_match and code_match:
                            position = p
                            break

                # 策略3: dept + title
                if not position and r['title']:
                    for p in positions:
                        if p.id in matched_ids:
                            continue
                        p_dept = (p.department or '').replace('\n', '')
                        if p_dept == r['department'] and (p.title or '') == r['title']:
                            position = p
                            break

                if position:
                    matched_ids.add(position.id)
                    matched += 1
                    if r['exam_weight_ratio']:
                        position.exam_weight_ratio = r['exam_weight_ratio']
                    if r['remark']:
                        position.remark = r['remark']
                else:
                    unmatched += 1

            total_matched += matched
            total_unmatched += unmatched

        await db.commit()
        print(f'  匹配: {total_matched}, 未匹配: {total_unmatched}')


async def main():
    for year in (2023, 2024):
        print(f'\n=== {year}年 ===')
        await merge_year(year)

    # 验证
    async with AsyncSessionLocal() as db:
        print('\n=== 验证 ===')
        for year in (2023, 2024, 2025):
            for field in ('exam_weight_ratio', 'remark'):
                col = getattr(Position, field)
                from sqlalchemy import func
                cnt = (await db.execute(select(func.count(Position.id)).where(
                    Position.year == year, Position.exam_type == '事业单位',
                    col.isnot(None), col != ''
                ))).scalar()
                total = (await db.execute(select(func.count(Position.id)).where(
                    Position.year == year, Position.exam_type == '事业单位'
                ))).scalar()
                print(f'  {year} {field}: {cnt}/{total}')


if __name__ == '__main__':
    asyncio.run(main())
