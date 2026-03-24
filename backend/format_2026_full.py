"""整理2026年江苏事业单位岗位表最全汇总.xls为标准Excel格式"""
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def clean(val):
    if val is None or val == '':
        return None
    s = str(val).strip().replace('\n', ' ')
    return s if s not in ('', '——', '-', '/') else None

def main():
    fp = '/Users/chaim/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_83do8l3kb13812_187a/msg/file/2026-03/江苏事业单位岗位表最全汇总(3).xls'
    wb_in = xlrd.open_workbook(fp)
    ws_in = wb_in.sheet_by_name('附件1')

    # 合并表头
    row1 = ws_in.row_values(1)
    row2 = ws_in.row_values(2)
    headers = []
    for i in range(ws_in.ncols):
        h1 = str(row1[i]).strip() if row1[i] else ''
        h2 = str(row2[i]).strip() if row2[i] else ''
        headers.append((h1 + h2).replace(' ', '').replace('\n', ''))

    # 列映射 (源索引 → 目标字段名)
    col_map = {
        0: ('city', '地市', 10),
        1: ('location', '区县', 12),
        2: ('supervising_dept', '主管部门', 22),
        4: ('department', '招录单位', 25),
        8: ('position_code', '岗位代码', 10),
        7: ('title', '岗位名称', 18),
        11: ('description', '岗位说明', 25),
        10: ('exam_category', '笔试科目', 14),
        12: ('position_level', '岗位等级', 12),
        6: ('funding_source', '经费来源', 12),
        13: ('recruitment_count', '招聘人数', 8),
        16: ('education', '学历', 14),
        17: ('major', '专业要求', 30),
        18: ('recruitment_target', '招聘对象', 12),
        19: ('other_requirements', '其他条件', 25),
        14: ('exam_ratio', '开考比例', 10),
        23: ('exam_weight_ratio', '考试形式占比', 20),
        25: ('remark', '其他说明', 20),
    }

    # 读取数据
    data = []
    for ri in range(3, ws_in.nrows):
        row = ws_in.row_values(ri)
        item = {}
        for src_idx, (field, _, _) in col_map.items():
            val = clean(row[src_idx]) if src_idx < len(row) else None
            if val:
                item[field] = val

        # 退役大学生士兵专项标注
        location = item.get('location', '')
        if '退役大学生士兵' in location and not item.get('major'):
            item['major'] = '退役大学生士兵专项（不限专业）'

        if item.get('department') or item.get('title'):
            data.append(item)

    print(f'解析数据: {len(data)} 条')

    # 导出Excel
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = '2026年江苏事业单位'

    # 输出列定义
    out_cols = [v for v in col_map.values()]

    # 样式
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    cell_align = Alignment(vertical='center')
    border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    # 写表头
    for ci, (field, header, width) in enumerate(out_cols, 1):
        cell = ws_out.cell(row=1, column=ci, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws_out.column_dimensions[get_column_letter(ci)].width = width

    ws_out.freeze_panes = 'A2'
    ws_out.auto_filter.ref = f'A1:{get_column_letter(len(out_cols))}1'

    # 写数据
    for ri, item in enumerate(data, 2):
        for ci, (field, _, _) in enumerate(out_cols, 1):
            val = item.get(field, '')
            cell = ws_out.cell(row=ri, column=ci, value=val)
            cell.alignment = cell_align
            cell.border = border

    out = '/Users/chaim/Desktop/2026年江苏事业单位总表_完整版.xlsx'
    wb_out.save(out)
    wb_out.close()
    print(f'已导出: {out}')


if __name__ == '__main__':
    main()
