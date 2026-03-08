"""合并进面分数据到 positions 表

进面分数据分布在132个文件中，格式各异。
通用策略：自动检测表头行和关键列，提取(单位, 岗位/代码, 笔试成绩)，
按(单位,岗位)分组计算 min/max 分数，匹配到 positions 表。
"""
import asyncio
import os
import re
import openpyxl
from sqlalchemy import select
from app.database import AsyncSessionLocal
from app.models.position import Position


# 城市目录 → 标准城市名
DIR_TO_CITY = {
    '01 省属': '省属', '02 南京市': '南京市', '03 镇江市': '镇江市',
    '04 扬州市': '扬州市', '05 常州市': '常州市', '06 泰州市': '泰州市',
    '07 无锡市': '无锡市', '08 苏州市': '苏州市', '09 南通市': '南通市',
    '10 盐城市': '盐城市', '11 宿迁市': '宿迁市', '12 淮安市': '淮安市',
    '13  连云港市': '连云港市', '14 徐州市': '徐州市',
}

# 表头关键词 → 列类型
HEADER_KEYWORDS = {
    'unit': ['招聘单位', '单位名称', '用人单位', '部门名称'],
    'position': ['招聘岗位', '岗位名称', '职位名称'],
    'position_code': ['岗位代码', '职位代码', '岗位\\n代码'],
    'score': ['笔试成绩', '笔试\\n成绩', '成绩', '合成\\n成绩'],
}


def detect_header_and_columns(ws) -> tuple[int, dict]:
    """自动检测表头行和列映射"""
    for ri in range(1, min(10, ws.max_row + 1)):
        row = []
        for cell in ws[ri]:
            row.append(str(cell.value).strip() if cell.value else '')

        # 判断是否是表头行：至少包含"成绩"或"笔试"关键词
        row_text = '|'.join(row)
        if '成绩' not in row_text and '笔试' not in row_text:
            continue

        col_map = {}
        for ci, val in enumerate(row):
            if not val:
                continue
            val_clean = val.replace('\n', '').replace(' ', '')
            for col_type, keywords in HEADER_KEYWORDS.items():
                if col_type in col_map:
                    continue
                for kw in keywords:
                    kw_clean = kw.replace('\n', '').replace(' ', '')
                    if kw_clean in val_clean:
                        col_map[col_type] = ci
                        break

        if 'score' in col_map:
            return ri, col_map

    return -1, {}


def parse_xlsx_scores(filepath: str) -> list[dict]:
    """解析 .xlsx 进面分文件"""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        return []

    results = []
    ws = wb.active
    header_row, col_map = detect_header_and_columns(ws)
    if header_row < 0:
        wb.close()
        return []

    score_col = col_map.get('score')
    unit_col = col_map.get('unit')
    pos_col = col_map.get('position')
    code_col = col_map.get('position_code')

    prev_unit = ''
    prev_pos = ''
    prev_code = ''

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue

        # 单位（可能合并单元格）
        if unit_col is not None and unit_col < len(row) and row[unit_col]:
            val = str(row[unit_col]).strip()
            if val and val not in ('None',):
                # 去掉 "主管部门-" 前缀
                if '-' in val:
                    val = val.split('-', 1)[-1].strip()
                prev_unit = val

        # 岗位名
        if pos_col is not None and pos_col < len(row) and row[pos_col]:
            val = str(row[pos_col]).strip()
            if val and val not in ('None',):
                prev_pos = val

        # 岗位代码
        if code_col is not None and code_col < len(row) and row[code_col]:
            val = str(row[code_col]).strip()
            if val and val not in ('None',):
                prev_code = val

        # 成绩
        score = None
        if score_col is not None and score_col < len(row) and row[score_col]:
            try:
                raw = str(row[score_col]).strip()
                if raw and raw not in ('-', '——', '', '缺考', '违纪', '免笔试'):
                    score = float(raw)
            except (ValueError, TypeError):
                pass

        if score is not None and (prev_unit or prev_pos):
            results.append({
                'unit': prev_unit,
                'position': prev_pos,
                'code': prev_code,
                'score': score,
            })

    wb.close()
    return results


def parse_xls_scores(filepath: str) -> list[dict]:
    """解析 .xls 进面分文件"""
    try:
        import xlrd
        wb = xlrd.open_workbook(filepath)
    except Exception:
        return []

    results = []
    sheet = wb.sheet_by_index(0)

    # 检测表头
    header_row = -1
    col_map = {}
    for ri in range(min(10, sheet.nrows)):
        row_text = '|'.join(str(sheet.cell_value(ri, ci)).strip() for ci in range(sheet.ncols))
        if '成绩' not in row_text and '笔试' not in row_text:
            continue
        for ci in range(sheet.ncols):
            val = str(sheet.cell_value(ri, ci)).strip().replace('\n', '').replace(' ', '')
            if not val:
                continue
            for col_type, keywords in HEADER_KEYWORDS.items():
                if col_type in col_map:
                    continue
                for kw in keywords:
                    if kw.replace('\n', '').replace(' ', '') in val:
                        col_map[col_type] = ci
                        break
        if 'score' in col_map:
            header_row = ri
            break

    if header_row < 0:
        return []

    score_col = col_map.get('score')
    unit_col = col_map.get('unit')
    pos_col = col_map.get('position')
    code_col = col_map.get('position_code')

    prev_unit = ''
    prev_pos = ''
    prev_code = ''

    for ri in range(header_row + 1, sheet.nrows):
        if unit_col is not None:
            val = str(sheet.cell_value(ri, unit_col)).strip()
            if val:
                if '-' in val:
                    val = val.split('-', 1)[-1].strip()
                prev_unit = val
        if pos_col is not None:
            val = str(sheet.cell_value(ri, pos_col)).strip()
            if val:
                prev_pos = val
        if code_col is not None:
            val = str(sheet.cell_value(ri, code_col)).strip()
            if val:
                prev_code = val

        score = None
        if score_col is not None:
            try:
                raw = str(sheet.cell_value(ri, score_col)).strip()
                if raw and raw not in ('-', '——', '', '缺考', '违纪', '免笔试', '0.0'):
                    score = float(raw)
            except (ValueError, TypeError):
                pass

        if score is not None and (prev_unit or prev_pos):
            results.append({
                'unit': prev_unit, 'position': prev_pos,
                'code': prev_code, 'score': score,
            })

    return results


def aggregate_scores(raw_scores: list[dict]) -> dict[tuple, dict]:
    """按(unit, position/code)分组，计算 min/max 分数"""
    groups = {}
    for s in raw_scores:
        key = (s['unit'], s['position'], s['code'])
        if key not in groups:
            groups[key] = {'min': s['score'], 'max': s['score'], 'count': 0}
        groups[key]['min'] = min(groups[key]['min'], s['score'])
        groups[key]['max'] = max(groups[key]['max'], s['score'])
        groups[key]['count'] += 1
    return groups


def codes_equal(a: str, b: str) -> bool:
    """比较代码，忽略前导零：'1' == '01' == '001'"""
    if not a or not b:
        return False
    a, b = str(a).strip(), str(b).strip()
    if a == b:
        return True
    try:
        return int(a) == int(b)
    except ValueError:
        return False


async def merge_scores(city: str, score_groups: dict):
    """将分数数据合并到 positions 表"""
    async with AsyncSessionLocal() as db:
        positions = (await db.execute(
            select(Position).where(
                Position.year == 2025,
                Position.exam_type == '事业单位',
                Position.city == city,
            )
        )).scalars().all()

        matched = 0
        unmatched = 0
        matched_ids = set()

        for (unit, pos, code), scores in score_groups.items():
            position = None

            # 策略1: 按 department + position_code 匹配（代码标准化比较）
            if code:
                for p in positions:
                    if p.id in matched_ids:
                        continue
                    code_match = codes_equal(p.position_code, code)
                    dept_match = (unit in (p.department or '')
                                 or (p.department or '') in unit) if unit else True
                    if code_match and dept_match:
                        position = p
                        break

            # 策略2: 按 department + title 匹配
            if not position and pos:
                for p in positions:
                    if p.id in matched_ids:
                        continue
                    title_match = (p.title or '') == pos
                    dept_match = (unit in (p.department or '')
                                 or (p.department or '') in unit) if unit else True
                    if title_match and dept_match:
                        position = p
                        break

            # 策略3: 只按 title 匹配（当 unit 为空时）
            if not position and pos and not unit:
                for p in positions:
                    if p.id in matched_ids:
                        continue
                    if (p.title or '') == pos:
                        position = p
                        break

            if position:
                position.min_interview_score = scores['min']
                position.max_interview_score = scores['max']
                matched_ids.add(position.id)
                matched += 1
            else:
                unmatched += 1

        await db.commit()
        return matched, unmatched


async def main():
    base = '/Users/chaim/Desktop/江苏事业单位2023~2025年进面分数线/2025年/2025江苏事业单位统考进面分汇总'

    total_matched = 0
    total_unmatched = 0
    total_files = 0
    total_scores = 0

    for city_dir in sorted(os.listdir(base)):
        city_path = os.path.join(base, city_dir)
        if not os.path.isdir(city_path):
            continue
        city = DIR_TO_CITY.get(city_dir)
        if not city:
            print(f'  跳过: {city_dir}')
            continue

        # 递归扫描该城市所有文件（包含子目录）
        all_scores = []
        file_count = 0
        for root, dirs, files in os.walk(city_path):
            for fname in sorted(files):
                fp = os.path.join(root, fname)
                if fname.endswith('.xlsx'):
                    scores = parse_xlsx_scores(fp)
                elif fname.endswith('.xls'):
                    scores = parse_xls_scores(fp)
                else:
                    continue
                file_count += 1
                all_scores.extend(scores)

        if not all_scores:
            print(f'  {city:5s}: {file_count}个文件, 无有效分数数据')
            continue

        # 聚合并合并
        groups = aggregate_scores(all_scores)
        matched, unmatched = await merge_scores(city, groups)

        total_matched += matched
        total_unmatched += unmatched
        total_files += file_count
        total_scores += len(all_scores)

        pct = matched / (matched + unmatched) * 100 if (matched + unmatched) > 0 else 0
        print(f'  {city:5s}: {file_count:2d}文件, {len(all_scores):4d}条成绩, '
              f'{len(groups):3d}岗位 → 匹配 {matched:3d}, 未匹配 {unmatched:2d} ({pct:.0f}%)')

    print(f'\n===== 总计 =====')
    print(f'文件: {total_files}, 成绩条数: {total_scores}')
    print(f'匹配岗位: {total_matched}, 未匹配: {total_unmatched}')


if __name__ == '__main__':
    asyncio.run(main())
