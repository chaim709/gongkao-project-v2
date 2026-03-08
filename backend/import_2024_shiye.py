"""2024年江苏事业单位一体化导入脚本
Step 1: 导入职位表 (5,575行)
Step 2: 合并报名数据
Step 3: 合并进面分汇总表
"""
import asyncio
import re
import os
import json
import zipfile
import tempfile
import shutil
import openpyxl
import warnings
from lxml import etree
from collections import defaultdict
from sqlalchemy import select, func, delete
from app.database import AsyncSessionLocal
from app.models.position import Position

warnings.filterwarnings('ignore')

YEAR = 2024
EXAM_TYPE = '事业单位'

CITY_NORMALIZE = {
    '省属': '省属', '南京': '南京市', '南京市': '南京市',
    '苏州': '苏州市', '无锡': '无锡市', '常州': '常州市',
    '南通': '南通市', '徐州': '徐州市', '扬州': '扬州市',
    '镇江': '镇江市', '盐城': '盐城市', '泰州': '泰州市',
    '淮安': '淮安市', '连云港': '连云港市', '宿迁': '宿迁市',
}

INVALID = ('——', '-', '', '/', '未公示', '#VALUE!', '#N/A', 'N/A')
CODE_PATTERN = re.compile(r'\[([^\]]+)\]\s*$')


def normalize_city(raw):
    if not raw:
        return None
    raw = str(raw).replace('\n', '').strip()
    # 去掉"市"后查表
    key = raw.replace('市', '') if raw not in ('省属',) else raw
    return CITY_NORMALIZE.get(key, CITY_NORMALIZE.get(raw, raw + '市'))


def extract_code(text):
    if not text:
        return '', ''
    text = str(text).strip()
    m = CODE_PATTERN.search(text)
    if m:
        code = m.group(1).strip()
        name = text[:m.start()].strip()
        if '-' in name:
            name = name.split('-', 1)[1].strip()
        return name, code
    return text, ''


def fix_xlsx(src):
    """修复损坏的 xlsx 样式表"""
    tmp = tempfile.mkdtemp()
    fixed = os.path.join(tmp, 'fixed.xlsx')
    with zipfile.ZipFile(src, 'r') as zin:
        with zipfile.ZipFile(fixed, 'w') as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == 'xl/styles.xml':
                    tree = etree.fromstring(data)
                    ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                    fills = tree.find('.//ns:fills', ns)
                    if fills is not None:
                        for fill in fills.findall('ns:fill', ns):
                            if len(list(fill)) == 0:
                                fills.remove(fill)
                    data = etree.tostring(tree, xml_declaration=True, encoding='UTF-8', standalone=True)
                zout.writestr(item, data)
    return fixed, tmp


# ============================================================
# Step 1: 导入职位表
# ============================================================
async def import_positions(db):
    fp = '/Users/chaim/Desktop/江苏事业单位2023~2025年进面分数线/2024年/[职位表]2024年江苏事业单位统考职位表汇总.xlsx'
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb['Sheet1']

    # 2024 列映射
    COL_MAP = {
        0: 'city', 1: 'location', 2: 'supervising_dept',
        3: 'department', 4: 'funding_source', 5: 'position_code',
        6: 'title', 7: 'description', 8: 'exam_category',
        9: 'exam_ratio', 10: 'recruitment_count', 11: 'education',
        12: 'major', 13: 'other_requirements', 14: 'recruitment_target',
    }

    # 合并单元格 forward-fill: city, location, supervising_dept, department, funding_source
    FILL_COLS = {0, 1, 2, 3, 4}

    inserted = 0
    skipped = 0
    prev_values = {}

    for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        try:
            filled = list(row)
            for ci in FILL_COLS:
                if ci < len(filled):
                    if filled[ci] is not None and str(filled[ci]).strip() != '':
                        prev_values[ci] = filled[ci]
                    else:
                        filled[ci] = prev_values.get(ci)

            data = {}
            for ci, field in COL_MAP.items():
                if ci >= len(filled):
                    continue
                val = filled[ci]
                if field == 'recruitment_count':
                    try:
                        data[field] = int(val) if val else 1
                    except (ValueError, TypeError):
                        data[field] = 1
                else:
                    if val is not None:
                        s = str(val).strip().replace('\n', '')
                        data[field] = s if s else None
                    else:
                        data[field] = None

            data['city'] = normalize_city(data.get('city'))
            data['year'] = YEAR
            data['exam_type'] = EXAM_TYPE
            data['province'] = '江苏省'

            if not data.get('title'):
                if data.get('description'):
                    data['title'] = data['description'][:100]
                elif data.get('position_code'):
                    data['title'] = f"岗位{data['position_code']}"
                else:
                    skipped += 1
                    continue
            if not data.get('department'):
                skipped += 1
                continue

            db.add(Position(**data))
            inserted += 1
            if inserted % 500 == 0:
                await db.flush()
        except Exception as e:
            skipped += 1
            if skipped <= 3:
                print(f'  跳过行{ri+2}: {e}')

    await db.flush()
    wb.close()
    return inserted, skipped


# ============================================================
# Step 2: 合并报名数据
# ============================================================
async def merge_apply(db):
    src = '/Users/chaim/Desktop/江苏事业单位2023~2025年进面分数线/2024年/[报名数据]2024年江苏事业单位统考报名人数汇总.xlsx'
    fixed, tmp_dir = fix_xlsx(src)
    wb = openpyxl.load_workbook(fixed, data_only=True)

    SHEET_CITY = {
        '省属': '省属', '南京': '南京市', '常州': '常州市', '南通': '南通市',
        '镇江': '镇江市', '连云港': '连云港市', '徐州': '徐州市', '扬州': '扬州市',
        '盐城': '盐城市', '宿迁': '宿迁市', '淮安': '淮安市', '泰州': '泰州市',
        '无锡': '无锡市', '苏州': '苏州市', '苏州常熟': '苏州市', '昆山': '苏州市',
        '姑苏': '苏州市', '太仓': '苏州市',
    }

    total_matched = 0
    total_unmatched = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        city = SHEET_CITY.get(sheet_name)
        if not city:
            continue

        # 解析报名数据
        apply_rows = []
        if sheet_name in ('苏州常熟', '昆山', '太仓'):
            # 格式: 序号, 招聘单位, 岗位类别, 岗位名称, 要求, 岗位代码, 招录人数, 报名人数, 审核通过
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[1]:
                    continue
                dept = str(row[1]).strip() if row[1] else ''
                pos = str(row[3]).strip() if row[3] else ''
                code = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                # 审核通过(已缴费) = 最后一个有效列
                apply_val = row[9] if len(row) > 9 and row[9] else (row[8] if len(row) > 8 else None)
                try:
                    apply_count = int(apply_val) if apply_val else None
                except (ValueError, TypeError):
                    apply_count = None
                if dept and apply_count is not None:
                    apply_rows.append({'dept_name': dept, 'dept_code': '', 'pos_name': pos, 'pos_code': code, 'apply_count': apply_count})
        elif sheet_name == '姑苏':
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[1]:
                    continue
                dept = str(row[1]).strip()
                pos = str(row[3]).strip() if row[3] else ''
                code = str(row[4]).strip() if row[4] else ''
                apply_val = row[7] if len(row) > 7 else None
                try:
                    apply_count = int(apply_val) if apply_val else None
                except (ValueError, TypeError):
                    apply_count = None
                if dept and apply_count is not None:
                    apply_rows.append({'dept_name': dept, 'dept_code': '', 'pos_name': pos, 'pos_code': code, 'apply_count': apply_count})
        elif sheet_name == '苏州':
            # 职位所属地区(0), 部门名称[code](1), 职位名称[code](2), ..., 报名成功(7)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[1]:
                    continue
                dn, dc = extract_code(row[1])
                pn, pc = extract_code(row[2])
                try:
                    ac = int(row[7]) if row[7] else None
                except (ValueError, TypeError):
                    ac = None
                if dn and ac is not None:
                    apply_rows.append({'dept_name': dn, 'dept_code': dc, 'pos_name': pn, 'pos_code': pc, 'apply_count': ac})
        else:
            # 标准格式: 部门名称[code](0), 职位名称[code](1), ..., 报名成功人数(6)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                dn, dc = extract_code(row[0])
                pn, pc = extract_code(row[1])
                try:
                    ac = int(row[6]) if row[6] else None
                except (ValueError, TypeError):
                    ac = None
                if dn and ac is not None:
                    apply_rows.append({'dept_name': dn, 'dept_code': dc, 'pos_name': pn, 'pos_code': pc, 'apply_count': ac})

        if not apply_rows:
            continue

        # 匹配到 DB
        positions = (await db.execute(
            select(Position).where(
                Position.year == YEAR, Position.exam_type == EXAM_TYPE, Position.city == city
            )
        )).scalars().all()

        by_name = {}
        for p in positions:
            by_name.setdefault((p.department or '', p.title or ''), []).append(p)

        matched = 0
        unmatched = 0
        matched_ids = set()

        for r in apply_rows:
            position = None
            pos_name = r['pos_name']
            # 宿迁等城市的 title 有编号前缀如 "19-城建规划岗"，去掉
            pos_name_clean = re.sub(r'^\d+-', '', pos_name)

            # 名称匹配
            candidates = by_name.get((r['dept_name'], pos_name), [])
            if not candidates:
                candidates = by_name.get((r['dept_name'], pos_name_clean), [])
            for c in candidates:
                if c.id not in matched_ids:
                    position = c
                    break
            # 模糊匹配
            if not position:
                for p in positions:
                    if p.id in matched_ids:
                        continue
                    dept_match = (r['dept_name'] in (p.department or '') or (p.department or '') in r['dept_name'])
                    title_match = (p.title or '') in (pos_name, pos_name_clean)
                    if dept_match and title_match:
                        position = p
                        break

            if position:
                position.apply_count = r['apply_count']
                if position.recruitment_count and position.recruitment_count > 0:
                    position.competition_ratio = round(r['apply_count'] / position.recruitment_count, 1)
                matched_ids.add(position.id)
                matched += 1
            else:
                unmatched += 1

        total_matched += matched
        total_unmatched += unmatched
        pct = matched / (matched + unmatched) * 100 if (matched + unmatched) > 0 else 0
        print(f'  {sheet_name:8s} ({city:5s}): 匹配 {matched:4d}, 未匹配 {unmatched:3d} ({pct:.0f}%)')

    await db.flush()
    wb.close()
    shutil.rmtree(tmp_dir)
    return total_matched, total_unmatched


# ============================================================
# Step 3: 合并进面分汇总表
# ============================================================
async def merge_scores(db):
    fp = '/Users/chaim/Desktop/江苏事业单位2023~2025年进面分数线/2024年/2024年进面分汇总表.xlsx'
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb.active

    # 按城市分组
    score_data = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        city, unit, pos, code = row[0], row[1], row[2], str(row[3] or '')
        min_s, max_s = row[4], row[5]
        if city and (min_s is not None or max_s is not None):
            city = normalize_city(city)
            score_data[city].append({
                'unit': unit or '', 'pos': pos or '', 'code': code,
                'min': min_s, 'max': max_s,
            })
    wb.close()

    total_matched = 0
    total_unmatched = 0

    for city, rows in sorted(score_data.items()):
        positions = (await db.execute(
            select(Position).where(
                Position.year == YEAR, Position.exam_type == EXAM_TYPE, Position.city == city
            )
        )).scalars().all()

        matched = 0
        unmatched = 0
        matched_ids = set()

        for r in rows:
            position = None
            # 策略1: dept + code
            if r['code']:
                for p in positions:
                    if p.id in matched_ids:
                        continue
                    try:
                        code_match = int(r['code']) == int(p.position_code or 0)
                    except (ValueError, TypeError):
                        code_match = r['code'] == (p.position_code or '')
                    dept_match = (r['unit'] in (p.department or '') or (p.department or '') in r['unit']) if r['unit'] else True
                    if code_match and dept_match:
                        position = p
                        break
            # 策略2: dept + title
            if not position and r['pos']:
                for p in positions:
                    if p.id in matched_ids:
                        continue
                    dept_match = (r['unit'] in (p.department or '') or (p.department or '') in r['unit']) if r['unit'] else True
                    title_match = (p.title or '') == r['pos']
                    if dept_match and title_match:
                        position = p
                        break

            if position:
                if r['min'] is not None:
                    try:
                        position.min_interview_score = float(r['min'])
                    except (ValueError, TypeError):
                        pass
                if r['max'] is not None:
                    try:
                        position.max_interview_score = float(r['max'])
                    except (ValueError, TypeError):
                        pass
                matched_ids.add(position.id)
                matched += 1
            else:
                unmatched += 1

        total_matched += matched
        total_unmatched += unmatched

    await db.flush()
    return total_matched, total_unmatched


# ============================================================
# Main
# ============================================================
async def main():
    async with AsyncSessionLocal() as db:
        # 删除旧数据
        old = (await db.execute(
            select(func.count(Position.id)).where(Position.year == YEAR, Position.exam_type == EXAM_TYPE)
        )).scalar()
        if old:
            print(f'删除旧 {YEAR} 事业单位数据: {old} 条')
            await db.execute(delete(Position).where(Position.year == YEAR, Position.exam_type == EXAM_TYPE))
            await db.flush()

        # Step 1
        print('\n=== Step 1: 导入职位表 ===')
        ins, skip = await import_positions(db)
        print(f'  插入: {ins}, 跳过: {skip}')

        # Step 2
        print('\n=== Step 2: 合并报名数据 ===')
        m, u = await merge_apply(db)
        print(f'  总匹配: {m}, 未匹配: {u}')

        # Step 3
        print('\n=== Step 3: 合并进面分 ===')
        m2, u2 = await merge_scores(db)
        print(f'  总匹配: {m2}, 未匹配: {u2}')

        await db.commit()

        # 验证
        total = (await db.execute(select(func.count(Position.id)).where(
            Position.year == YEAR, Position.exam_type == EXAM_TYPE))).scalar()
        wa = (await db.execute(select(func.count(Position.id)).where(
            Position.year == YEAR, Position.exam_type == EXAM_TYPE, Position.apply_count.isnot(None)))).scalar()
        ws_min = (await db.execute(select(func.count(Position.id)).where(
            Position.year == YEAR, Position.exam_type == EXAM_TYPE, Position.min_interview_score.isnot(None)))).scalar()
        ws_max = (await db.execute(select(func.count(Position.id)).where(
            Position.year == YEAR, Position.exam_type == EXAM_TYPE, Position.max_interview_score.isnot(None)))).scalar()
        print(f'\n=== 最终结果 ===')
        print(f'总岗位:       {total}')
        print(f'有报名人数:   {wa} ({wa/total*100:.0f}%)')
        print(f'有最低进面分: {ws_min} ({ws_min/total*100:.0f}%)')
        print(f'有最高进面分: {ws_max} ({ws_max/total*100:.0f}%)')

asyncio.run(main())
