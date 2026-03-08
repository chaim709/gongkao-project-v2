"""2025年江苏事业单位岗位表导入脚本

数据源: [职位表]2025年江苏事业单位统一公开招聘岗位表.xlsx
- 14个Sheet（每个城市一个）
- 双行表头（Row 1标题, Row 2-3表头），Row 4开始数据
- Col 19 有地区数据，Col 17/18 分数列为空（分数在单独文件）
"""
import asyncio
import openpyxl
from sqlalchemy import select, func, delete
from app.database import AsyncSessionLocal
from app.models.position import Position


# 列索引 → 字段映射
COLUMN_MAP = {
    0: 'supervising_dept',    # 部门名称
    # 1: 主管部门代码 - 全空，跳过
    2: 'department',          # 单位名称
    3: 'department_code',     # 单位代码
    4: 'funding_source',      # 经费来源
    5: 'title',               # 岗位名称
    6: 'position_code',       # 岗位代码
    7: 'exam_category',       # 岗位类别
    8: 'recruitment_count',   # 招聘人数
    9: 'exam_ratio',          # 开考比例
    10: 'education',          # 学历
    11: 'major',              # 专业
    12: 'other_requirements', # 其他条件
    13: 'recruitment_target', # 招聘对象
    14: 'description',        # 考试形式和所占比例
    # 15: 其他说明 - 不入库
    # 16: 政策咨询电话 - 不入库
    # 17: 最高分 - 全空
    # 18: 最低分 - 全空
    19: 'location',           # 地区（区县）
}

# 向下填充列 - 只有 Col 0 (部门名称) 在苏州有合并单元格
# 安全起见对所有sheet都做，1%空值的不会被误填
FILL_DOWN_COLS = {0}

# Sheet名 → 标准城市名
SHEET_TO_CITY = {
    '省属': '省属',
    '南京': '南京市', '苏州': '苏州市', '无锡': '无锡市',
    '常州': '常州市', '南通': '南通市', '徐州': '徐州市',
    '扬州': '扬州市', '镇江': '镇江市', '盐城': '盐城市',
    '泰州': '泰州市', '淮安': '淮安市', '连云港': '连云港市',
    '宿迁': '宿迁市',
}

# 无效值集合
INVALID_VALUES = {'——', '-', '', '/', '未公示', '#VALUE!', '#N/A', 'N/A', '最高分', '最低分', '地区'}


async def main():
    fp = '/Users/chaim/Desktop/江苏事业单位2023~2025年进面分数线/2025年/[职位表]2025年江苏事业单位统一公开招聘岗位表.xlsx'
    year = 2025

    wb = openpyxl.load_workbook(fp, data_only=True)
    print(f'打开文件成功，共 {len(wb.sheetnames)} 个 Sheet')

    async with AsyncSessionLocal() as db:
        # Step 1: 删除旧数据
        old_count = (await db.execute(
            select(func.count(Position.id)).where(
                Position.year == year, Position.exam_type == '事业单位'
            )
        )).scalar()
        print(f'\n删除旧数据: {old_count} 条')
        await db.execute(
            delete(Position).where(
                Position.year == year, Position.exam_type == '事业单位'
            )
        )
        await db.flush()

        # Step 2: 逐 Sheet 导入
        total_inserted = 0
        total_skipped = 0
        sheet_stats = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            city = SHEET_TO_CITY.get(sheet_name, sheet_name + '市')
            inserted = 0
            skipped = 0
            prev_values = {}

            for ri, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
                try:
                    filled_row = list(row)

                    # 向下填充合并单元格
                    for ci in FILL_DOWN_COLS:
                        if ci < len(filled_row):
                            if filled_row[ci] is not None and str(filled_row[ci]).strip() != '':
                                prev_values[ci] = filled_row[ci]
                            else:
                                filled_row[ci] = prev_values.get(ci)

                    row_data = {}
                    for ci, field in COLUMN_MAP.items():
                        if ci >= len(filled_row):
                            continue
                        val = filled_row[ci]

                        if field == 'recruitment_count':
                            try:
                                row_data[field] = int(val) if val and str(val).strip() not in INVALID_VALUES else 1
                            except (ValueError, TypeError):
                                row_data[field] = 1
                        else:
                            if val is not None:
                                s = str(val).strip()
                                row_data[field] = s if s and s not in INVALID_VALUES else None
                            else:
                                row_data[field] = None

                    # 固定字段
                    row_data['year'] = year
                    row_data['exam_type'] = '事业单位'
                    row_data['province'] = '江苏省'
                    row_data['city'] = city

                    # 跳过无效行
                    if not row_data.get('title') and not row_data.get('department'):
                        skipped += 1
                        continue
                    if not row_data.get('title'):
                        if row_data.get('position_code'):
                            row_data['title'] = f"岗位{row_data['position_code']}"
                        else:
                            skipped += 1
                            continue

                    # 清理换行符
                    for field in ('title', 'department', 'supervising_dept',
                                  'education', 'major', 'other_requirements',
                                  'recruitment_target', 'description', 'exam_ratio',
                                  'exam_category', 'funding_source'):
                        if row_data.get(field):
                            row_data[field] = row_data[field].replace('\n', '').replace('\r', '')

                    db.add(Position(**row_data))
                    inserted += 1

                    if inserted % 500 == 0:
                        await db.flush()

                except Exception as e:
                    skipped += 1
                    if skipped <= 3:
                        print(f'  跳过 {sheet_name} 行{ri+4}: {e}')

            await db.flush()
            total_inserted += inserted
            total_skipped += skipped
            sheet_stats.append((sheet_name, city, inserted, skipped))
            print(f'  {sheet_name:6s} ({city:5s}): 插入 {inserted:4d}, 跳过 {skipped}')

        await db.commit()

    wb.close()

    print(f'\n===== 导入完成 =====')
    print(f'总插入: {total_inserted}')
    print(f'总跳过: {total_skipped}')
    print(f'预期行数: 5945')

    if total_inserted != 5945:
        print(f'⚠️  差异: {5945 - total_inserted} 行')


if __name__ == '__main__':
    asyncio.run(main())
