"""2023年江苏事业单位一体化导入脚本
Step 1: 导入职位表 (5,701行)
Step 2: 合并报名数据 (16个Sheet)
Step 3: 合并进面分 (102个xlsx/xls + pdf/docx)
"""
import asyncio
import re
import os
import openpyxl
import warnings
from collections import defaultdict
from sqlalchemy import select, func, delete
from app.database import AsyncSessionLocal
from app.models.position import Position

warnings.filterwarnings('ignore')

YEAR = 2023
EXAM_TYPE = '事业单位'

CITY_NORMALIZE = {
    '省属': '省属', '海事局': '省属', '地震局': '省属',
    '南京': '南京市', '南京市': '南京市',
    '苏州': '苏州市', '无锡': '无锡市', '常州': '常州市',
    '南通': '南通市', '徐州': '徐州市', '扬州': '扬州市',
    '镇江': '镇江市', '盐城': '盐城市', '泰州': '泰州市',
    '淮安': '淮安市', '连云港': '连云港市', '宿迁': '宿迁市',
}

CODE_PATTERN = re.compile(r'\[([^\]]+)\]\s*$')

HEADER_KEYWORDS = {
    'unit': ['招聘单位', '单位名称', '用人单位', '部门名称'],
    'position': ['招聘岗位', '岗位名称', '职位名称', '岗位代码岗位名称'],
    'position_code': ['岗位代码', '职位代码', '岗位\n代码'],
    'score': ['笔试成绩', '笔试\n成绩', '成绩', '合成\n成绩'],
}

# 从 "(code)name" 格式中提取代码和名称
PAREN_CODE = re.compile(r'^\(([^)]+)\)\s*(.+)')


def extract_paren_code(text):
    """从 '(code)name' 格式提取 → (name, code)"""
    if not text:
        return '', ''
    text = str(text).strip()
    m = PAREN_CODE.match(text)
    if m:
        return m.group(2).strip(), m.group(1).strip()
    return text, ''


def normalize_city(raw):
    if not raw:
        return None
    raw = str(raw).replace('\n', '').strip()
    return CITY_NORMALIZE.get(raw, CITY_NORMALIZE.get(raw.replace('市', ''), raw + '市'))


def extract_code(text):
    if not text:
        return '', ''
    text = str(text).strip()
    m = CODE_PATTERN.search(text)
    if m:
        code = m.group(1).strip()
        name = text[:m.start()].strip()
        if '-' in name:
            name = name.split('-', 1)[1].strip()
        return name, code
    return text, ''


# ============================================================
# Step 1: 导入职位表
# ============================================================
async def import_positions(db):
    fp = '/Users/chaim/Desktop/江苏事业单位2023~2025年进面分数线/2023年/[职位表]2023年江苏事业单位统考职位表汇总.xlsx'
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb['Sheet1']

    # 2023列映射 (与2024相同)
    COL_MAP = {
        0: 'city', 1: 'location', 2: 'supervising_dept',
        3: 'department', 4: 'funding_source', 5: 'position_code',
        6: 'title', 7: 'description', 8: 'exam_category',
        9: 'exam_ratio', 10: 'recruitment_count', 11: 'education',
        12: 'major', 13: 'other_requirements', 14: 'recruitment_target',
    }

    FILL_COLS = {0, 1, 2, 3, 4}
    inserted = 0
    skipped = 0
    prev_values = {}
    prev_city = None

    for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        try:
            filled = list(row)

            # 城市切换时重置 forward-fill 状态
            current_city = filled[0]
            if current_city is not None and str(current_city).strip() != '':
                if prev_city is not None and str(current_city).strip() != str(prev_city).strip():
                    prev_values = {0: current_city}
                prev_city = current_city

            for ci in FILL_COLS:
                if ci < len(filled):
                    if filled[ci] is not None and str(filled[ci]).strip() != '':
                        prev_values[ci] = filled[ci]
                    else:
                        filled[ci] = prev_values.get(ci)

            data = {}
            for ci, field in COL_MAP.items():
                if ci >= len(filled):
                    continue
                val = filled[ci]
                if field == 'recruitment_count':
                    try:
                        data[field] = int(val) if val else 1
                    except (ValueError, TypeError):
                        data[field] = 1
                else:
                    if val is not None:
                        s = str(val).strip().replace('\n', '')
                        data[field] = s if s else None
                    else:
                        data[field] = None

            data['city'] = normalize_city(data.get('city'))
            data['year'] = YEAR
            data['exam_type'] = EXAM_TYPE
            data['province'] = '江苏省'

            if not data.get('title'):
                if data.get('description'):
                    data['title'] = data['description'][:100]
                elif data.get('position_code'):
                    data['title'] = f"岗位{data['position_code']}"
                else:
                    skipped += 1
                    continue
            if not data.get('department'):
                # 海事局/地震局的 department 为空，用 supervising_dept 作为 fallback
                if data.get('supervising_dept'):
                    data['department'] = data['supervising_dept']
                elif data.get('city') == '省属' and data.get('title'):
                    # 地震局等无部门信息的，构造部门名
                    data['department'] = '江苏省地震局'
                else:
                    skipped += 1
                    continue

            db.add(Position(**data))
            inserted += 1
            if inserted % 500 == 0:
                await db.flush()
        except Exception as e:
            skipped += 1
            if skipped <= 3:
                print(f'  跳过行{ri+2}: {e}')

    await db.flush()
    wb.close()
    return inserted, skipped


# ============================================================
# Step 2: 合并报名数据
# ============================================================
async def merge_apply(db):
    src = '/Users/chaim/Desktop/江苏事业单位2023~2025年进面分数线/2023年/[报名数据]2023年江苏事业单位统考报名人数汇总.xlsx'
    wb = openpyxl.load_workbook(src, data_only=True)

    # Sheet名 → (city, apply_col, data_start, 是否苏州特殊格式)
    SHEET_CONFIG = {
        '省属事业单位': ('省属', 4, 2, False),
        '海事局': ('省属', 4, 2, False),
        '地震局': ('省属', 4, 2, False),
        '南京': ('南京市', 6, 2, False),
        '苏州': ('苏州市', 9, 3, True),  # 特殊格式
        '无锡': ('无锡市', 4, 2, False),
        '常州': ('常州市', 6, 2, False),
        '镇江': ('镇江市', 5, 2, False),
        '南通': ('南通市', 4, 2, False),
        '扬州': ('扬州市', 5, 2, False),
        '泰州': ('泰州市', 4, 2, False),
        '淮安': ('淮安市', 4, 2, False),
        '盐城': ('盐城市', 4, 2, False),
        '徐州': ('徐州市', 4, 2, False),
        '宿迁': ('宿迁市', 4, 2, False),
        '连云港': ('连云港市', 4, 2, False),
    }

    total_matched = 0
    total_unmatched = 0

    for sheet_name in wb.sheetnames:
        config = SHEET_CONFIG.get(sheet_name)
        if not config:
            continue
        city, apply_col, data_start, is_suzhou = config
        ws = wb[sheet_name]

        apply_rows = []
        if is_suzhou:
            # 苏州: 地区(0), 单位代码(1), 单位名称(2), 岗位代码(3), 岗位名称(4), ..., 已缴费(9)
            for row in ws.iter_rows(min_row=data_start, values_only=True):
                if not row or not row[2]:
                    continue
                dept = str(row[2]).strip() if row[2] else ''
                pos = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                code = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                apply_val = row[apply_col] if len(row) > apply_col else None
                try:
                    apply_count = int(apply_val) if apply_val else None
                except (ValueError, TypeError):
                    apply_count = None
                if dept and apply_count is not None:
                    apply_rows.append({
                        'dept_name': dept, 'dept_code': '',
                        'pos_name': pos, 'pos_code': code,
                        'apply_count': apply_count,
                    })
        else:
            # 标准格式: dept[code](0), pos[code](1), ..., 报名成功(apply_col)
            for row in ws.iter_rows(min_row=data_start, values_only=True):
                if not row or not row[0]:
                    continue
                dn, dc = extract_code(row[0])
                pn, pc = extract_code(row[1])
                apply_val = row[apply_col] if len(row) > apply_col else None
                try:
                    apply_count = int(apply_val) if apply_val else None
                except (ValueError, TypeError):
                    apply_count = None
                if dn and apply_count is not None:
                    apply_rows.append({
                        'dept_name': dn, 'dept_code': dc,
                        'pos_name': pn, 'pos_code': pc,
                        'apply_count': apply_count,
                    })

        if not apply_rows:
            continue

        # 匹配到 DB
        positions = (await db.execute(
            select(Position).where(
                Position.year == YEAR, Position.exam_type == EXAM_TYPE, Position.city == city
            )
        )).scalars().all()

        by_name = {}
        for p in positions:
            by_name.setdefault((p.department or '', p.title or ''), []).append(p)

        # 建代码索引: (dept_name, pos_code) → [positions]
        by_code = {}
        for p in positions:
            if p.position_code:
                by_code.setdefault((p.department or '', p.position_code or ''), []).append(p)

        matched = 0
        unmatched = 0
        matched_ids = set()

        for r in apply_rows:
            position = None
            pos_name = r['pos_name']

            # 策略1: 名称匹配 (dept + title)
            candidates = by_name.get((r['dept_name'], pos_name), [])
            for c in candidates:
                if c.id not in matched_ids:
                    position = c
                    break

            # 策略2: 代码匹配 (dept + pos_code)
            if not position and r['pos_code']:
                candidates = by_code.get((r['dept_name'], r['pos_code']), [])
                for c in candidates:
                    if c.id not in matched_ids:
                        position = c
                        break

            # 策略3: 模糊名称匹配
            if not position:
                for p in positions:
                    if p.id in matched_ids:
                        continue
                    dept_match = (r['dept_name'] in (p.department or '')
                                 or (p.department or '') in r['dept_name'])
                    title_match = (p.title or '') == pos_name
                    if dept_match and title_match:
                        position = p
                        break

            # 策略4: 模糊dept + code匹配
            if not position and r['pos_code']:
                for p in positions:
                    if p.id in matched_ids:
                        continue
                    dept_match = (r['dept_name'] in (p.department or '')
                                 or (p.department or '') in r['dept_name'])
                    code_match = codes_equal(p.position_code, r['pos_code'])
                    if dept_match and code_match:
                        position = p
                        break

            if position:
                position.apply_count = r['apply_count']
                if position.recruitment_count and position.recruitment_count > 0:
                    position.competition_ratio = round(r['apply_count'] / position.recruitment_count, 1)
                matched_ids.add(position.id)
                matched += 1
            else:
                unmatched += 1

        total_matched += matched
        total_unmatched += unmatched
        pct = matched / (matched + unmatched) * 100 if (matched + unmatched) > 0 else 0
        print(f'  {sheet_name:8s} ({city:5s}): 匹配 {matched:4d}, 未匹配 {unmatched:3d} ({pct:.0f}%)')

    await db.flush()
    wb.close()
    return total_matched, total_unmatched


# ============================================================
# Step 3: 合并进面分 (从散文件解析)
# ============================================================
DIR_TO_CITY = {
    '01 省属': '省属', '02 南京市': '南京市', '02 南京市(1)': '南京市',
    '03 镇江市': '镇江市', '04 扬州市': '扬州市', '05 常州市': '常州市',
    '06 泰州市': '泰州市', '07 无锡市': '无锡市', '08 苏州市': '苏州市',
    '09 南通市': '南通市', '10 盐城市': '盐城市', '11 宿迁市': '宿迁市',
    '12 淮安市': '淮安市', '13 连云港': '连云港市', '14 徐州市': '徐州市',
}

INVALID_SCORES = ('-', '——', '', '缺考', '违纪', '免笔试', '0.0', 'None', '/')


def detect_header_and_columns(ws) -> tuple:
    for ri in range(1, min(10, ws.max_row + 1)):
        row = []
        for cell in ws[ri]:
            row.append(str(cell.value).strip() if cell.value else '')

        row_text = '|'.join(row)
        if '成绩' not in row_text and '笔试' not in row_text:
            continue

        col_map = {}
        for ci, val in enumerate(row):
            if not val:
                continue
            val_clean = val.replace('\n', '').replace(' ', '')
            for col_type, keywords in HEADER_KEYWORDS.items():
                if col_type in col_map:
                    continue
                for kw in keywords:
                    kw_clean = kw.replace('\n', '').replace(' ', '')
                    if kw_clean in val_clean:
                        col_map[col_type] = ci
                        break

        if 'score' in col_map:
            return ri, col_map

    return -1, {}


def parse_xlsx_scores(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        return []

    results = []
    ws = wb.active
    header_row, col_map = detect_header_and_columns(ws)
    if header_row < 0:
        wb.close()
        return []

    score_col = col_map.get('score')
    unit_col = col_map.get('unit')
    pos_col = col_map.get('position')
    code_col = col_map.get('position_code')

    prev_unit = ''
    prev_pos = ''
    prev_code = ''

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue
        if unit_col is not None and unit_col < len(row) and row[unit_col]:
            val = str(row[unit_col]).strip()
            if val and val != 'None':
                # 处理 "(code)name" 格式
                name, _ = extract_paren_code(val)
                if '-' in name:
                    name = name.split('-', 1)[-1].strip()
                prev_unit = name
        if pos_col is not None and pos_col < len(row) and row[pos_col]:
            val = str(row[pos_col]).strip()
            if val and val != 'None':
                name, code = extract_paren_code(val)
                prev_pos = name
                if code:
                    prev_code = code
        if code_col is not None and code_col < len(row) and row[code_col]:
            val = str(row[code_col]).strip()
            if val and val != 'None':
                prev_code = val

        score = None
        if score_col is not None and score_col < len(row) and row[score_col]:
            try:
                raw = str(row[score_col]).strip()
                if raw and raw not in INVALID_SCORES:
                    score = float(raw)
            except (ValueError, TypeError):
                pass

        if score is not None and (prev_unit or prev_pos):
            results.append({
                'unit': prev_unit, 'position': prev_pos,
                'code': prev_code, 'score': score,
            })

    wb.close()
    return results


def parse_xls_scores(filepath):
    try:
        import xlrd
        wb = xlrd.open_workbook(filepath)
    except Exception:
        return []

    results = []
    sheet = wb.sheet_by_index(0)

    header_row = -1
    col_map = {}
    for ri in range(min(10, sheet.nrows)):
        row_text = '|'.join(str(sheet.cell_value(ri, ci)).strip() for ci in range(sheet.ncols))
        if '成绩' not in row_text and '笔试' not in row_text:
            continue
        for ci in range(sheet.ncols):
            val = str(sheet.cell_value(ri, ci)).strip().replace('\n', '').replace(' ', '')
            if not val:
                continue
            for col_type, keywords in HEADER_KEYWORDS.items():
                if col_type in col_map:
                    continue
                for kw in keywords:
                    if kw.replace('\n', '').replace(' ', '') in val:
                        col_map[col_type] = ci
                        break
        if 'score' in col_map:
            header_row = ri
            break

    if header_row < 0:
        return []

    score_col = col_map.get('score')
    unit_col = col_map.get('unit')
    pos_col = col_map.get('position')
    code_col = col_map.get('position_code')

    prev_unit = ''
    prev_pos = ''
    prev_code = ''

    for ri in range(header_row + 1, sheet.nrows):
        if unit_col is not None:
            val = str(sheet.cell_value(ri, unit_col)).strip()
            if val:
                name, _ = extract_paren_code(val)
                if '-' in name:
                    name = name.split('-', 1)[-1].strip()
                prev_unit = name
        if pos_col is not None:
            val = str(sheet.cell_value(ri, pos_col)).strip()
            if val:
                name, code = extract_paren_code(val)
                prev_pos = name
                if code:
                    prev_code = code
        if code_col is not None:
            val = str(sheet.cell_value(ri, code_col)).strip()
            if val:
                prev_code = val

        score = None
        if score_col is not None:
            try:
                raw = str(sheet.cell_value(ri, score_col)).strip()
                if raw and raw not in INVALID_SCORES:
                    score = float(raw)
            except (ValueError, TypeError):
                pass

        if score is not None and (prev_unit or prev_pos):
            results.append({
                'unit': prev_unit, 'position': prev_pos,
                'code': prev_code, 'score': score,
            })

    return results


def parse_pdf_scores(filepath):
    try:
        import pdfplumber
    except ImportError:
        return []

    results = []
    try:
        pdf = pdfplumber.open(filepath)
    except Exception:
        return []

    for page in pdf.pages:
        tables = page.extract_tables()
        for table in tables:
            if not table:
                continue
            # 找表头
            header_row_idx = -1
            col_map = {}
            for ri, row in enumerate(table):
                row_text = '|'.join(str(c or '').strip() for c in row)
                if '成绩' not in row_text and '笔试' not in row_text:
                    continue
                for ci, val in enumerate(row):
                    if not val:
                        continue
                    val_clean = str(val).replace('\n', '').replace(' ', '').strip()
                    for col_type, keywords in HEADER_KEYWORDS.items():
                        if col_type in col_map:
                            continue
                        for kw in keywords:
                            if kw.replace('\n', '').replace(' ', '') in val_clean:
                                col_map[col_type] = ci
                                break
                if 'score' in col_map:
                    header_row_idx = ri
                    break

            if header_row_idx < 0:
                continue

            score_col = col_map.get('score')
            unit_col = col_map.get('unit')
            pos_col = col_map.get('position')
            code_col = col_map.get('position_code')

            prev_unit = ''
            prev_pos = ''
            prev_code = ''

            for row in table[header_row_idx + 1:]:
                if unit_col is not None and unit_col < len(row) and row[unit_col]:
                    val = str(row[unit_col]).strip()
                    if val:
                        name, _ = extract_paren_code(val)
                        if '-' in name:
                            name = name.split('-', 1)[-1].strip()
                        prev_unit = name
                if pos_col is not None and pos_col < len(row) and row[pos_col]:
                    val = str(row[pos_col]).strip()
                    if val:
                        name, code = extract_paren_code(val)
                        prev_pos = name
                        if code:
                            prev_code = code
                if code_col is not None and code_col < len(row) and row[code_col]:
                    val = str(row[code_col]).strip()
                    if val:
                        prev_code = val

                score = None
                if score_col is not None and score_col < len(row) and row[score_col]:
                    try:
                        raw = str(row[score_col]).strip()
                        if raw and raw not in INVALID_SCORES:
                            score = float(raw)
                    except (ValueError, TypeError):
                        pass

                if score is not None and (prev_unit or prev_pos):
                    results.append({
                        'unit': prev_unit, 'position': prev_pos,
                        'code': prev_code, 'score': score,
                    })

    pdf.close()
    return results


def parse_docx_scores(filepath):
    try:
        from docx import Document
    except ImportError:
        return []

    results = []
    try:
        doc = Document(filepath)
    except Exception:
        return []

    for table in doc.tables:
        rows_data = []
        for row in table.rows:
            rows_data.append([cell.text.strip() for cell in row.cells])

        if not rows_data:
            continue

        header_row_idx = -1
        col_map = {}
        for ri, row in enumerate(rows_data):
            row_text = '|'.join(row)
            if '成绩' not in row_text and '笔试' not in row_text:
                continue
            for ci, val in enumerate(row):
                if not val:
                    continue
                val_clean = val.replace('\n', '').replace(' ', '')
                for col_type, keywords in HEADER_KEYWORDS.items():
                    if col_type in col_map:
                        continue
                    for kw in keywords:
                        if kw.replace('\n', '').replace(' ', '') in val_clean:
                            col_map[col_type] = ci
                            break
            if 'score' in col_map:
                header_row_idx = ri
                break

        if header_row_idx < 0:
            continue

        score_col = col_map.get('score')
        unit_col = col_map.get('unit')
        pos_col = col_map.get('position')
        code_col = col_map.get('position_code')

        prev_unit = ''
        prev_pos = ''
        prev_code = ''

        for row in rows_data[header_row_idx + 1:]:
            if unit_col is not None and unit_col < len(row) and row[unit_col]:
                val = row[unit_col]
                if val:
                    name, _ = extract_paren_code(val)
                    if '-' in name:
                        name = name.split('-', 1)[-1].strip()
                    prev_unit = name
            if pos_col is not None and pos_col < len(row) and row[pos_col]:
                val = row[pos_col]
                if val:
                    name, code = extract_paren_code(val)
                    prev_pos = name
                    if code:
                        prev_code = code
            if code_col is not None and code_col < len(row) and row[code_col]:
                val = row[code_col]
                if val:
                    prev_code = val

            score = None
            if score_col is not None and score_col < len(row) and row[score_col]:
                try:
                    raw = row[score_col].strip()
                    if raw and raw not in INVALID_SCORES:
                        score = float(raw)
                except (ValueError, TypeError):
                    pass

            if score is not None and (prev_unit or prev_pos):
                results.append({
                    'unit': prev_unit, 'position': prev_pos,
                    'code': prev_code, 'score': score,
                })

    return results


def aggregate_scores(raw_scores):
    groups = {}
    for s in raw_scores:
        key = (s['unit'], s['position'], s['code'])
        if key not in groups:
            groups[key] = {'min': s['score'], 'max': s['score'], 'count': 0}
        groups[key]['min'] = min(groups[key]['min'], s['score'])
        groups[key]['max'] = max(groups[key]['max'], s['score'])
        groups[key]['count'] += 1
    return groups


def codes_equal(a, b):
    if not a or not b:
        return False
    a, b = str(a).strip(), str(b).strip()
    if a == b:
        return True
    try:
        return int(a) == int(b)
    except ValueError:
        return False


async def merge_scores(db):
    base = '/Users/chaim/Desktop/江苏事业单位2023~2025年进面分数线/2023年/2023江苏事业单位统考进面分汇总'

    total_matched = 0
    total_unmatched = 0
    total_files = 0
    total_raw = 0

    for city_dir in sorted(os.listdir(base)):
        city_path = os.path.join(base, city_dir)
        if not os.path.isdir(city_path):
            continue
        city = DIR_TO_CITY.get(city_dir)
        if not city:
            print(f'  跳过: {city_dir}')
            continue

        all_scores = []
        file_count = 0
        for root, dirs, files in os.walk(city_path):
            for fname in sorted(files):
                fp = os.path.join(root, fname)
                if fname.endswith('.xlsx'):
                    scores = parse_xlsx_scores(fp)
                elif fname.endswith('.xls'):
                    scores = parse_xls_scores(fp)
                elif fname.endswith('.pdf'):
                    scores = parse_pdf_scores(fp)
                elif fname.endswith(('.docx', '.doc')):
                    scores = parse_docx_scores(fp)
                else:
                    continue
                file_count += 1
                all_scores.extend(scores)

        if not all_scores:
            if file_count > 0:
                print(f'  {city:5s}: {file_count}个文件, 无有效分数数据')
            continue

        # 聚合
        groups = aggregate_scores(all_scores)

        # 匹配
        positions = (await db.execute(
            select(Position).where(
                Position.year == YEAR, Position.exam_type == EXAM_TYPE, Position.city == city
            )
        )).scalars().all()

        matched = 0
        unmatched = 0
        matched_ids = set()

        for (unit, pos, code), scores in groups.items():
            position = None

            # 策略1: dept + code
            if code:
                for p in positions:
                    if p.id in matched_ids:
                        continue
                    code_match = codes_equal(p.position_code, code)
                    dept_match = (unit in (p.department or '')
                                 or (p.department or '') in unit) if unit else True
                    if code_match and dept_match:
                        position = p
                        break

            # 策略2: dept + title
            if not position and pos:
                for p in positions:
                    if p.id in matched_ids:
                        continue
                    title_match = (p.title or '') == pos
                    dept_match = (unit in (p.department or '')
                                 or (p.department or '') in unit) if unit else True
                    if title_match and dept_match:
                        position = p
                        break

            # 策略3: 只按 title
            if not position and pos and not unit:
                for p in positions:
                    if p.id in matched_ids:
                        continue
                    if (p.title or '') == pos:
                        position = p
                        break

            if position:
                position.min_interview_score = scores['min']
                position.max_interview_score = scores['max']
                matched_ids.add(position.id)
                matched += 1
            else:
                unmatched += 1

        total_matched += matched
        total_unmatched += unmatched
        total_files += file_count
        total_raw += len(all_scores)

        pct = matched / (matched + unmatched) * 100 if (matched + unmatched) > 0 else 0
        print(f'  {city:5s}: {file_count:2d}文件, {len(all_scores):4d}条成绩, '
              f'{len(groups):3d}岗位 → 匹配 {matched:3d}, 未匹配 {unmatched:2d} ({pct:.0f}%)')

    await db.flush()
    return total_matched, total_unmatched, total_files, total_raw


# ============================================================
# Main
# ============================================================
async def main():
    async with AsyncSessionLocal() as db:
        # 删除旧数据
        old = (await db.execute(
            select(func.count(Position.id)).where(Position.year == YEAR, Position.exam_type == EXAM_TYPE)
        )).scalar()
        if old:
            print(f'删除旧 {YEAR} 事业单位数据: {old} 条')
            await db.execute(delete(Position).where(Position.year == YEAR, Position.exam_type == EXAM_TYPE))
            await db.flush()

        # Step 1
        print('\n=== Step 1: 导入职位表 ===')
        ins, skip = await import_positions(db)
        print(f'  插入: {ins}, 跳过: {skip}')

        # Step 2
        print('\n=== Step 2: 合并报名数据 ===')
        m, u = await merge_apply(db)
        print(f'  总匹配: {m}, 未匹配: {u}')

        # Step 3
        print('\n=== Step 3: 合并进面分 ===')
        m2, u2, fc, rc = await merge_scores(db)
        print(f'  文件: {fc}, 成绩条数: {rc}')
        print(f'  总匹配: {m2}, 未匹配: {u2}')

        await db.commit()

        # 验证
        total = (await db.execute(select(func.count(Position.id)).where(
            Position.year == YEAR, Position.exam_type == EXAM_TYPE))).scalar()
        wa = (await db.execute(select(func.count(Position.id)).where(
            Position.year == YEAR, Position.exam_type == EXAM_TYPE, Position.apply_count.isnot(None)))).scalar()
        ws_min = (await db.execute(select(func.count(Position.id)).where(
            Position.year == YEAR, Position.exam_type == EXAM_TYPE, Position.min_interview_score.isnot(None)))).scalar()
        ws_max = (await db.execute(select(func.count(Position.id)).where(
            Position.year == YEAR, Position.exam_type == EXAM_TYPE, Position.max_interview_score.isnot(None)))).scalar()
        print(f'\n=== 最终结果 ===')
        print(f'总岗位:       {total}')
        print(f'有报名人数:   {wa} ({wa/total*100:.0f}%)')
        print(f'有最低进面分: {ws_min} ({ws_min/total*100:.0f}%)')
        print(f'有最高进面分: {ws_max} ({ws_max/total*100:.0f}%)')

asyncio.run(main())
